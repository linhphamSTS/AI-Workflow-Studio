#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""Render a cloud cost estimation workbook, one sheet per environment.

    python build_cost.py --sizing sizing.json --out "Cost Estimation - Acme.xlsx"

`sizing.json` holds the priced design:

    {
      "project": "Acme", "region": "...", "priced_on": "2026-07-28",
      "hours_per_month": 730,
      "environments": [{"key": "prod", "name": "Production", "note": "..."}],
      "prices": {"m7g.large": {"usd": 0.1, "unit": "hour", "source": "MEC1-BoxUsage:..."}},
      "lines": [{"layer": "Compute", "service": "EKS worker nodes",
                 "config": "why it is this size",
                 "price_key": "m7g.large", "hourly": true,
                 "reservable": true,      // vendor sells reserved/committed pricing for it
                 "stoppable": true,       // it has a stopped state, so a schedule works
                 "qty": {"prod": 4, "stg": 2, "dev": 3}}],
      "already_applied": [{"name": "...", "why": "..."}],
      "levers": [{"name": "...", "what": "...", "scope": "non-prod hourly",
                  "pct": 0.643, "how": "..."}],
      "reserved": [{"term": "1 year", "option": "No Upfront", "pct": 0.338, "note": "..."}],
      "third_party": [{"item": "...", "what": "...", "basis": "...", "treatment": "..."}]
    }

One sheet per environment because a reader pricing staging should not have to read past
production to find it, and because a challenged sizing gets argued one environment at a
time.

Every money cell is a formula over the quantity, so a corrected quantity flows through to
the cover without anyone recomputing anything.
"""
import argparse
import json
import os
import sys

from openpyxl import Workbook
from openpyxl.utils import get_column_letter

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
from xlsx_style import (  # noqa: E402
    A_CENTER, A_LEFT_CENTER, A_LEFT_TOP, C_EFFORT, C_HEADER, C_HEADER_TOT, C_L1, C_L2,
    C_LEAF, C_TOTAL, MONEY, PCT, autofit_row_heights, banner, notes, style,
)

ENV_W = {'A': 5, 'B': 16, 'C': 32, 'D': 50, 'E': 17, 'F': 13, 'G': 9, 'H': 9,
         'I': 15, 'J': 16, 'K': 54}
OPT_W = {'A': 5, 'B': 40, 'C': 62, 'D': 17, 'E': 10, 'F': 14, 'G': 14, 'H': 15, 'I': 62}
TP_W = {'A': 5, 'B': 34, 'C': 64, 'D': 30, 'E': 70}
SRC_W = {'A': 5, 'B': 26, 'C': 14, 'D': 18, 'E': 86}
COV_W = {'A': 44, 'B': 18, 'C': 18, 'D': 18, 'E': 18, 'F': 18}


def head(ws, row, labels, rgb_for=None):
    for col, h in enumerate(labels, start=1):
        ws.cell(row, col, h)
        style(ws.cell(row, col), rgb=(rgb_for(col) if rgb_for else C_HEADER),
              bold=True, white=True)


def build_env(ws, spec, env):
    ws.sheet_view.showGridLines = False
    for letter, w in ENV_W.items():
        ws.column_dimensions[letter].width = w
    span = len(ENV_W)
    hours = spec.get('hours_per_month', 730)

    banner(ws, 1, '%s  -  %s' % (env['name'].upper(), spec['region']), C_HEADER, span, 13)
    ws.row_dimensions[1].height = 24
    banner(ws, 2, '%s List prices extracted %s; every row names the source it came from. '
                  'Quantity is the only input on this sheet.'
           % (env.get('note', ''), spec['priced_on']), C_EFFORT, span)

    r = 4
    head(ws, r, ['#', 'Layer', 'Service', 'Configuration', 'Unit', 'Unit price\n(USD)',
                 'Hrs /\nmonth', 'Qty', 'USD / month', 'USD / year', 'Source'],
         rgb_for=lambda c: C_HEADER_TOT if c in (9, 10) else C_HEADER)
    hdr = r
    r += 1

    # `run` is every hourly row. The two narrower lists exist because a lever is only
    # honest over the rows the vendor actually sells it for: a control plane, a load
    # balancer and a NAT gateway cannot be reserved, and a managed cache, search or
    # streaming tier has no stopped state, so a nightly schedule buys nothing there.
    # Reservability and stoppability are INDEPENDENT facts, not one derived from the
    # other: a managed database can usually be stopped on a tier that cannot be reserved.
    layers, keys, run, reservable, stoppable = {}, {}, [], [], []
    n = 0
    for line in spec['lines']:
        qty = (line.get('qty') or {}).get(env['key'], 0)
        if not qty:
            continue
        n += 1
        p = spec['prices'][line['price_key']]
        h = hours if line.get('hourly') else 1
        for col, v in enumerate([n, line['layer'], line['service'], line.get('config', ''),
                                 p['unit'], p['usd'], h, qty, None, None,
                                 p.get('source', '')], start=1):
            style(ws.cell(r, col, v), rgb=C_LEAF,
                  align=A_LEFT_TOP if col in (2, 3, 4, 5, 11) else A_CENTER)
        ws.cell(r, 9, '=$F%d*$G%d*$H%d' % (r, r, r))
        ws.cell(r, 10, '=I%d*12' % r)
        for col in (6, 9, 10):
            ws.cell(r, col).number_format = MONEY
        layers.setdefault(line['layer'], []).append(r)
        keys.setdefault(line['price_key'], []).append(r)
        if line.get('hourly'):
            run.append(r)
            if line.get('reservable', True):
                reservable.append(r)
            if line.get('stoppable', True):
                stoppable.append(r)
        r += 1

    r += 1
    banner(ws, r, 'SUBTOTAL BY LAYER', C_HEADER, span)
    r += 1
    first = r
    for layer in sorted(layers):
        ws.cell(r, 2, layer)
        style(ws.cell(r, 2), rgb=C_L2, bold=True, align=A_LEFT_CENTER)
        for col in list(range(1, 9)) + [11]:
            style(ws.cell(r, col), rgb=C_L2)
        for col in (9, 10):
            L = get_column_letter(col)
            ws.cell(r, col, '=' + '+'.join('%s%d' % (L, x) for x in layers[layer]))
            style(ws.cell(r, col), rgb=C_L2, bold=True)
            ws.cell(r, col).number_format = MONEY
        r += 1
    last = r - 1

    ws.cell(r, 2, 'TOTAL  %s' % env['name'].upper())
    style(ws.cell(r, 2), rgb=C_TOTAL, bold=True, align=A_LEFT_CENTER)
    for col in list(range(1, 9)) + [11]:
        style(ws.cell(r, col), rgb=C_TOTAL)
    for col in (9, 10):
        L = get_column_letter(col)
        ws.cell(r, col, '=SUM(%s%d:%s%d)' % (L, first, L, last))
        style(ws.cell(r, col), rgb=C_TOTAL, bold=True)
        ws.cell(r, col).number_format = MONEY

    autofit_row_heights(ws, ENV_W, skip_rows={1, hdr})
    ws.row_dimensions[hdr].height = 40
    return {'sheet': ws.title, 'total_row': r, 'run': run,
            'reservable': reservable, 'stoppable': stoppable, 'keys': keys,
            'key': env['key'], 'name': env['name']}


def build_optimisation(ws, spec, envs):
    ws.sheet_view.showGridLines = False
    for letter, w in OPT_W.items():
        ws.column_dimensions[letter].width = w
    span = len(OPT_W)
    prod, nonprod = envs[0], envs[1:]

    def basis(scope):
        # A scope naming an instrument resolves to the rows eligible for it, not to every
        # hourly row. Claiming a lever over an ineligible line is the defect this exists to
        # prevent, and it is invisible in the finished workbook.
        if scope == 'non-prod hourly':
            t = [f"'{e['sheet']}'!I{x}" for e in nonprod for x in e['run']]
        elif scope == 'non-prod stoppable':
            t = [f"'{e['sheet']}'!I{x}" for e in nonprod for x in e['stoppable']]
        elif scope == 'prod reservable':
            t = [f"'{prod['sheet']}'!I{x}" for x in prod['reservable']]
        elif scope == 'all reservable':
            t = [f"'{e['sheet']}'!I{x}" for e in envs for x in e['reservable']]
        elif scope == 'prod hourly':
            t = [f"'{prod['sheet']}'!I{x}" for x in prod['run']]
        elif scope in ('none', '', None):
            t = []
        else:
            t = [f"'{e['sheet']}'!I{x}" for e in envs for x in e['keys'].get(scope, [])]
        return ('=' + '+'.join(t)) if t else 0

    banner(ws, 1, 'COST OPTIMISATION  -  what each lever removes, and from what',
           C_HEADER, span, 13)
    ws.row_dimensions[1].height = 24
    banner(ws, 2, 'Nothing here is netted off the environment sheets. Each lever names the '
                  'rows it acts on so the reader can choose which to bank.', C_EFFORT, span)

    r = 4
    if spec.get('already_applied'):
        banner(ws, r, 'ALREADY APPLIED IN THE SIZING  -  not available to take again',
               C_L1, span)
        style(ws.cell(r, 1), rgb=C_L1, bold=True, align=A_LEFT_CENTER)
        for col in range(2, span + 1):
            style(ws.cell(r, col), rgb=C_L1)
        r += 1
        for i, item in enumerate(spec['already_applied'], start=1):
            for col, v in enumerate([i, item['name'], item['why'], 'in the sizing'], start=1):
                style(ws.cell(r, col, v), rgb=C_LEAF,
                      align=A_LEFT_TOP if col in (2, 3) else A_CENTER)
            ws.merge_cells(start_row=r, start_column=4, end_row=r, end_column=span)
            r += 1
        r += 1

    banner(ws, r, 'AVAILABLE LEVERS  -  each one still costs something to take', C_HEADER, span)
    r += 1
    head(ws, r, ['#', 'Lever', 'What it does', 'Applies to', 'Saving %', 'Basis USD/mo',
                 'Saving USD/mo', 'Saving USD/yr', 'How it is done'])
    hdr = r
    r += 1
    for i, lv in enumerate(spec.get('levers', []), start=1):
        for col, v in enumerate([i, lv['name'], lv['what'], lv.get('scope', 'none'),
                                 lv.get('pct', 0), basis(lv.get('scope')), None, None,
                                 lv.get('how', '')], start=1):
            style(ws.cell(r, col, v), rgb=C_LEAF,
                  align=A_LEFT_TOP if col in (2, 3, 9) else A_CENTER)
        ws.cell(r, 7, '=$F%d*$E%d' % (r, r))
        ws.cell(r, 8, '=G%d*12' % r)
        ws.cell(r, 5).number_format = PCT
        for col in (6, 7, 8):
            ws.cell(r, col).number_format = MONEY
        r += 1

    if spec.get('reserved'):
        r += 1
        banner(ws, r, 'COMMITMENT DISCOUNTS  -  read from the price list, applied to '
                      'steady-state production only', C_HEADER, span)
        r += 1
        head(ws, r, ['#', 'Term', 'Purchase option', 'Discount', 'Prod base USD/mo',
                     'Committed USD/mo', 'Committed USD/yr', 'Saving USD/yr', 'Note'])
        hdr2 = r
        r += 1
        for i, ri in enumerate(spec['reserved'], start=1):
            for col, v in enumerate([i, ri['term'], ri['option'], ri['pct'],
                                     basis('prod reservable'), None, None, None,
                                     ri.get('note', '')], start=1):
                style(ws.cell(r, col, v), rgb=C_LEAF,
                      align=A_LEFT_TOP if col == 9 else A_CENTER)
            ws.cell(r, 6, '=$E%d*(1-$D%d)' % (r, r))
            ws.cell(r, 7, '=F%d*12' % r)
            ws.cell(r, 8, '=($E%d-$F%d)*12' % (r, r))
            ws.cell(r, 4).number_format = PCT
            for col in (5, 6, 7, 8):
                ws.cell(r, col).number_format = MONEY
            r += 1
        r += 1
        r = notes(ws, r, span, 'WHY THE COMMITMENT ONLY APPLIES TO PRODUCTION', [
            'A commitment bills every hour of the term whether the resource runs or not. '
            'Committing to an environment that is deliberately switched off at night pays '
            'for exactly the hours the schedule just saved, so the two cancel out.',
            'Non-production takes the schedule, production takes the commitment. The two '
            'savings are then additive because they act on different sheets.',
        ])
    else:
        hdr2 = hdr

    autofit_row_heights(ws, OPT_W, skip_rows={1, hdr, hdr2})
    for h in {hdr, hdr2}:
        ws.row_dimensions[h].height = 32


def build_third_party(ws, spec):
    ws.sheet_view.showGridLines = False
    for letter, w in TP_W.items():
        ws.column_dimensions[letter].width = w
    span = len(TP_W)
    banner(ws, 1, 'THIRD-PARTY AND LICENCE COST  -  not the cloud bill, and not in the '
                  'environment totals', C_HEADER, span, 13)
    ws.row_dimensions[1].height = 24
    banner(ws, 2, 'Listed with their pricing basis even where the amount depends on a '
                  'decision the client has not taken. Leaving them out makes the platform '
                  'look cheaper than it is.', C_EFFORT, span)
    r = 4
    head(ws, r, ['#', 'Item', 'What it is for', 'Pricing basis', 'Treatment'])
    hdr = r
    r += 1
    for i, tp in enumerate(spec.get('third_party', []), start=1):
        for col, v in enumerate([i, tp['item'], tp['what'], tp['basis'], tp['treatment']],
                                start=1):
            style(ws.cell(r, col, v), rgb=C_LEAF,
                  align=A_CENTER if col == 1 else A_LEFT_TOP)
        r += 1
    autofit_row_heights(ws, TP_W, skip_rows={1, hdr})
    ws.row_dimensions[hdr].height = 24


def build_sources(ws, spec):
    ws.sheet_view.showGridLines = False
    for letter, w in SRC_W.items():
        ws.column_dimensions[letter].width = w
    span = len(SRC_W)
    banner(ws, 1, 'PRICE SOURCES  -  every unit price and where it came from',
           C_HEADER, span, 13)
    ws.row_dimensions[1].height = 24
    banner(ws, 2, 'Extracted %s for %s. LIST prices: they exclude any enterprise agreement, '
                  'credits or private pricing the client holds. Re-extract before the number '
                  'goes out.' % (spec['priced_on'], spec['region']), C_EFFORT, span)
    r = 4
    head(ws, r, ['#', 'Key', 'USD', 'Unit', 'Source'])
    hdr = r
    r += 1
    for i, (key, p) in enumerate(sorted(spec['prices'].items()), start=1):
        for col, v in enumerate([i, key, p['usd'], p['unit'], p.get('source', '')], start=1):
            style(ws.cell(r, col, v), rgb=C_LEAF,
                  align=A_LEFT_TOP if col in (2, 4, 5) else A_CENTER)
        ws.cell(r, 3).number_format = '#,##0.000000'
        r += 1
    autofit_row_heights(ws, SRC_W, skip_rows={1, hdr})
    ws.row_dimensions[hdr].height = 24


def build_cover(ws, spec, envs):
    ws.sheet_view.showGridLines = False
    for letter, w in COV_W.items():
        ws.column_dimensions[letter].width = w
    span = len(COV_W)
    banner(ws, 1, '%s  -  INFRASTRUCTURE COST ESTIMATION' % spec['project'].upper(),
           C_HEADER, span, 14)
    ws.row_dimensions[1].height = 26
    banner(ws, 2, '%d environment(s) on %s' % (len(envs), spec['region']), C_EFFORT, span)

    r = 4
    for label, value in [('Region', spec['region']),
                         ('Prices extracted', '%s, vendor price list (list prices)'
                          % spec['priced_on']),
                         ('Environments', ', '.join(e['name'] for e in envs)),
                         ('Excluded', spec.get('excluded', 'Support plan, tax, marketplace '
                                               'subscriptions, and the third-party licences '
                                               'on their own sheet.'))]:
        ws.cell(r, 1, label)
        style(ws.cell(r, 1), rgb=C_L2, bold=True, align=A_LEFT_CENTER)
        ws.cell(r, 2, value)
        for col in range(2, span + 1):
            style(ws.cell(r, col), align=A_LEFT_CENTER)
        ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=span)
        r += 1

    r += 1
    head(ws, r, ['Environment', 'USD / month', 'USD / year', 'Share', '', ''],
         rgb_for=lambda c: C_HEADER_TOT if c in (2, 3) else C_HEADER)
    ws.merge_cells(start_row=r, start_column=5, end_row=r, end_column=span)
    r += 1
    first, total_at = r, r + len(envs)
    for e in envs:
        ws.cell(r, 1, e['name'])
        ws.cell(r, 2, "='%s'!I%d" % (e['sheet'], e['total_row']))
        ws.cell(r, 3, "='%s'!J%d" % (e['sheet'], e['total_row']))
        ws.cell(r, 4, '=IF($B$%d=0,0,B%d/$B$%d)' % (total_at, r, total_at))
        for col in range(1, span + 1):
            style(ws.cell(r, col), rgb=C_L1)
        style(ws.cell(r, 1), rgb=C_L1, bold=True, align=A_LEFT_CENTER)
        for col in (2, 3):
            ws.cell(r, col).number_format = MONEY
        ws.cell(r, 4).number_format = PCT
        ws.merge_cells(start_row=r, start_column=5, end_row=r, end_column=span)
        r += 1
    ws.cell(r, 1, 'TOTAL')
    for col in range(1, span + 1):
        style(ws.cell(r, col), rgb=C_TOTAL)
    style(ws.cell(r, 1), rgb=C_TOTAL, bold=True, align=A_LEFT_CENTER)
    for col in (2, 3):
        L = get_column_letter(col)
        # Name each environment row rather than sum the range. The range happens to hold only
        # environment rows today, so it is arithmetically right, but it is the same fragile
        # shape that tripled a delivered WBS cover the moment a reader inserted a subtotal
        # inside it. Named cells cannot be widened by accident.
        ws.cell(r, col, '=' + '+'.join('%s%d' % (L, x) for x in range(first, r)))
        style(ws.cell(r, col), rgb=C_TOTAL, bold=True)
        ws.cell(r, col).number_format = MONEY
    ws.merge_cells(start_row=r, start_column=5, end_row=r, end_column=span)
    r += 2

    notes(ws, r, span, 'HOW TO READ THIS WORKBOOK', [
        'One sheet per environment. Quantity is the only input; every money cell is a '
        'formula, so a challenged sizing is corrected in place and this cover follows.',
        'These figures are the un-optimised, on-demand list position. Optimisation lists '
        'each lever and the rows it acts on, and nothing is netted off, so every reduction '
        'stays a visible choice.',
        'Price Sources carries every unit price with the source it came from, so any line '
        'can be traced back rather than taken on trust.',
    ])
    autofit_row_heights(ws, COV_W, skip_rows={1})


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument('--sizing', required=True)
    ap.add_argument('--out', required=True)
    a = ap.parse_args()
    spec = json.load(open(a.sizing, encoding='utf-8'))

    missing = sorted({l['price_key'] for l in spec['lines']} - set(spec['prices']))
    if missing:
        sys.exit('! lines reference prices that are not in `prices`: %s' % ', '.join(missing))

    wb = Workbook()
    cover = wb.active
    cover.title = 'Cover'
    envs = [build_env(wb.create_sheet(e['name'][:31]), spec, e)
            for e in spec['environments']]
    build_optimisation(wb.create_sheet('Optimisation'), spec, envs)
    build_third_party(wb.create_sheet('Third-Party'), spec)
    build_sources(wb.create_sheet('Price Sources'), spec)
    build_cover(cover, spec, envs)
    wb.save(a.out)

    hours = spec.get('hours_per_month', 730)
    print('Written: %s' % a.out)
    print('Sheets : %s' % ', '.join(wb.sheetnames))
    print()
    grand = 0.0
    for e in spec['environments']:
        tot = sum(spec['prices'][l['price_key']]['usd']
                  * (hours if l.get('hourly') else 1)
                  * (l.get('qty') or {}).get(e['key'], 0) for l in spec['lines'])
        grand += tot
        print('  %-16s %14s /mo  %16s /yr'
              % (e['name'], format(tot, ',.2f'), format(tot * 12, ',.2f')))
    print('  %-16s %14s /mo  %16s /yr'
          % ('TOTAL', format(grand, ',.2f'), format(grand * 12, ',.2f')))


if __name__ == '__main__':
    main()
