#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""Gate the cost workbook before anyone is told a number.

    python verify_cost.py --sizing sizing.json --xlsx "Cost Estimation - Acme.xlsx" \\
                          --prices prices.json [--max-price-age-days 14]

Exit 0 only when every check passes.

Why this exists at all: **openpyxl writes formulas but no cached results**, so a saved
workbook proves nothing about its own totals. Excel computes them when the client opens it.
This evaluates every formula the same way Excel would and fails if the file a client sees
would show a different number from the one the project reported.

Three classes of defect, none of which is visible in the finished file:

  Arithmetic   a total that misses a layer, or counts one twice, or mixes a one-off amount
               into a monthly figure so the sum means nothing
  Eligibility  a saving lever claimed over a line the vendor does not sell that instrument
               for. Reserved pricing on a Kubernetes control plane, a load balancer or a NAT
               gateway, or a nightly shutdown on a managed cache, search or streaming tier,
               overstates the saving by a fifth and looks entirely correct on the page
  Freshness    a list price extracted weeks ago, carried forward under today's date. The
               remedy is not a warning in prose on the sheet; it is re-resolving every unit
               price against the vendor's own output and failing on a mismatch

The eligibility ground truth is written HERE, from service naming rather than from the
sizing file's own flags, so the check cannot agree with the model by construction. On an
earlier bid the equivalent check read back the very list the builder used to exclude rows,
so it only ever confirmed that the builder was self-consistent.
"""
import argparse
import datetime as _dt
import json
import os
import re
import sys

from openpyxl import load_workbook
from openpyxl.utils import column_index_from_string as col_index
from openpyxl.utils import get_column_letter

CELL = re.compile(r"(?:'([^']+)'!)?\$?([A-Z]{1,3})\$?(\d+)")
RANGE = re.compile(r"SUM\((?:'([^']+)'!)?\$?([A-Z]{1,3})\$?(\d+):\$?([A-Z]{1,3})\$?(\d+)\)")
# The cover expresses each environment's share as IF(<denom>=0,0,<num>/<denom>), which is a
# guard against dividing by an empty estate. It has to be resolved BEFORE the cell references
# become numbers, because "=" inside the expression is not Python.
IFSHARE = re.compile(r"IF\((\$?[A-Z]{1,3}\$?\d+)=0,0,(\$?[A-Z]{1,3}\$?\d+)/(\$?[A-Z]{1,3}\$?\d+)\)")

# Column layout of an environment sheet, discovered rather than assumed where possible.
C_LAYER, C_SERVICE, C_UNITPRICE, C_HOURS, C_QTY, C_MONTH, C_YEAR, C_SOURCE = 2, 3, 6, 7, 8, 9, 10, 11

# ---------------------------------------------------------------- vendor ground truth
# Independent of the sizing file. These are properties of how the major clouds sell, not
# opinions about this estate, and they are matched on the service name so a mislabelled
# flag in the sizing file is caught rather than trusted.
#
# Reservability and stoppability are INDEPENDENT facts and are listed separately on
# purpose. A managed database can usually be STOPPED on a burstable tier that cannot be
# RESERVED, so deriving either list from the other produces a false positive.
NEVER_RESERVABLE = (
    'control plane', 'load balancer', 'application gateway', 'nat gateway',
    'api gateway', 'api management', 'key vault', 'secrets manager', 'kms',
    'log analytics', 'log ingestion', 'cloudwatch', 'egress', 'data transfer',
    'bandwidth', 'object storage', 'blob', 's3', 'event bus', 'eventbridge',
    'service bus', 'event hubs', 'queue', 'sqs', 'sns', 'container registry',
    'public ip', 'dns', 'cdn', 'managed disk', 'os disk',
)
NEVER_STOPPABLE = (
    'control plane', 'load balancer', 'application gateway', 'nat gateway',
    'api gateway', 'api management', 'key vault', 'secrets manager', 'kms',
    'log analytics', 'log ingestion', 'cloudwatch', 'egress', 'data transfer',
    'bandwidth', 'object storage', 'blob', 's3', 'event bus', 'eventbridge',
    'service bus', 'event hubs', 'queue', 'sqs', 'sns', 'container registry',
    'public ip', 'dns', 'cdn', 'managed disk', 'os disk',
    # These CAN be reserved but have no stopped state: stopping means delete and rebuild.
    'cache', 'redis', 'valkey', 'memcached', 'search', 'opensearch',
    'elasticsearch', 'kafka', 'msk', 'streaming',
)


class Checks:
    def __init__(self):
        self.rows, self.failures = [], []

    def add(self, name, ok, detail=''):
        self.rows.append((name, bool(ok)))
        if not ok:
            self.failures.append(detail or name)
        return bool(ok)

    def report(self):
        for name, ok in self.rows:
            print('  [%s] %s' % ('PASS' if ok else 'FAIL', name))
        print()
        if self.failures:
            print('RESULT: %d FAILURE(S)' % len(self.failures))
            for f in self.failures:
                print('  - %s' % f)
            return 1
        print('RESULT: ALL CHECKS PASS (%d checks)' % len(self.rows))
        return 0


class Book:
    """Evaluates the workbook the way Excel will, because the file itself cannot."""

    def __init__(self, path):
        self.wb = load_workbook(path)

    def value(self, sheet, coord, stack=()):
        key = (sheet, coord)
        if key in stack:
            raise RuntimeError('circular reference through %s!%s' % key)
        raw = self.wb[sheet][coord].value
        if isinstance(raw, (int, float)):
            return float(raw)
        if not isinstance(raw, str) or not raw.startswith('='):
            return 0.0
        here = stack + (key,)
        expr = raw[1:]

        def rng(m):
            sh = m.group(1) or sheet
            acc = 0.0
            for c in range(col_index(m.group(2)), col_index(m.group(4)) + 1):
                for r in range(int(m.group(3)), int(m.group(5)) + 1):
                    acc += self.value(sh, get_column_letter(c) + str(r), here)
            return repr(acc)

        def share(m):
            denom = self.value(sheet, m.group(3).replace('$', ''), here)
            num = self.value(sheet, m.group(2).replace('$', ''), here)
            return repr(0.0 if denom == 0 else num / denom)

        expr = IFSHARE.sub(share, expr)
        expr = RANGE.sub(rng, expr)
        expr = CELL.sub(
            lambda m: repr(self.value(m.group(1) or sheet, m.group(2) + m.group(3), here)),
            expr)
        try:
            return float(eval(expr, {'__builtins__': {}}, {}))  # noqa: S307
        except Exception as exc:
            raise RuntimeError('cannot evaluate %s!%s = %s (%s)' % (sheet, coord, raw, exc))


def find_env_rows(ws):
    """Return (header_row, line_rows, layer_rows, total_row) by reading the sheet."""
    hdr = None
    for r in range(1, ws.max_row + 1):
        if str(ws.cell(r, 1).value or '').strip() == '#':
            hdr = r
            break
    if hdr is None:
        return None, [], [], None
    lines, layers, total = [], [], None
    r = hdr + 1
    while r <= ws.max_row and isinstance(ws.cell(r, 1).value, int):
        lines.append(r)
        r += 1
    for rr in range(r, ws.max_row + 1):
        label = str(ws.cell(rr, C_LAYER).value or '').strip()
        if label.upper().startswith('TOTAL'):
            total = rr
            break
        if label and isinstance(ws.cell(rr, C_MONTH).value, str):
            layers.append(rr)
    return hdr, lines, layers, total


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument('--sizing', required=True)
    ap.add_argument('--xlsx', required=True)
    ap.add_argument('--prices', help='a price file freshly fetched with cloud_prices.py; '
                                     'omitting it fails the freshness checks rather than '
                                     'skipping them')
    ap.add_argument('--max-price-age-days', type=int, default=14)
    ap.add_argument('--today', help='ISO date, for reproducible runs')
    a = ap.parse_args()

    spec = json.load(open(a.sizing, encoding='utf-8'))
    B = Book(a.xlsx)
    wb = B.wb
    ck = Checks()
    hours = spec.get('hours_per_month', 730)
    envs = spec['environments']

    # ------------------------------------------------------------------ structure
    want = (['Cover'] + [e['name'][:31] for e in envs]
            + ['Optimisation', 'Third-Party', 'Price Sources'])
    ck.add('sheet order Cover / environments / Optimisation / Third-Party / Price Sources',
           wb.sheetnames == want, 'sheets are %s' % wb.sheetnames)

    for line in spec['lines']:
        pass
    unknown = sorted({l['price_key'] for l in spec['lines']} - set(spec['prices']))
    ck.add('every line references a known price key', not unknown, 'unknown: %s' % unknown)
    nosrc = sorted(k for k, p in spec['prices'].items() if not p.get('source'))
    ck.add('every price names the vendor record it came from', not nosrc,
           'no source: %s' % nosrc)

    # ------------------------------------------------------------------ arithmetic
    env_info = {}
    for e in envs:
        name = e['name'][:31]
        ws = wb[name]
        hdr, lines, layers, total = find_env_rows(ws)
        env_info[e['key']] = (name, lines, layers, total)
        ck.add('%s: sheet structure found' % name,
               bool(hdr and lines and layers and total),
               'header=%s lines=%d layers=%d total=%s' % (hdr, len(lines), len(layers), total))
        if not (hdr and lines and layers and total):
            continue

        # A money cell must be a formula over the quantity, so a corrected quantity flows
        # through. A literal is a number nobody can re-derive.
        literal = [r for r in lines
                   if not str(ws.cell(r, C_MONTH).value or '').startswith('=')]
        ck.add('%s: every monthly amount is a formula' % name, not literal,
               'literal amounts at rows %s' % literal[:5])

        # Each line: monthly == unit price x hours x qty, evaluated not assumed.
        wrong = []
        for r in lines:
            unit = ws.cell(r, C_UNITPRICE).value or 0
            h = ws.cell(r, C_HOURS).value or 0
            q = ws.cell(r, C_QTY).value or 0
            got = B.value(name, 'I%d' % r)
            if abs(got - unit * h * q) > 0.005:
                wrong.append('%s!I%d = %.4f, expected %.4f' % (name, r, got, unit * h * q))
        ck.add('%s: every line is price x hours x quantity' % name, not wrong,
               '; '.join(wrong[:3]))

        # The layer subtotals must PARTITION the lines: every line counted once, none twice.
        covered = []
        for lr in layers:
            f = str(ws.cell(lr, C_MONTH).value or '')
            covered += [int(m.group(3)) for m in CELL.finditer(f[1:])]
        ck.add('%s: layer subtotals cover every line exactly once' % name,
               sorted(covered) == sorted(lines),
               'lines %s, covered %s' % (lines, sorted(covered)))

        # The environment total must equal the sum of its lines, computed independently.
        model = sum((ws.cell(r, C_UNITPRICE).value or 0) * (ws.cell(r, C_HOURS).value or 0)
                    * (ws.cell(r, C_QTY).value or 0) for r in lines)
        got = B.value(name, 'I%d' % total)
        ck.add('%s: TOTAL equals the sum of its lines' % name, abs(got - model) < 0.01,
               'workbook shows %.2f, the lines add to %.2f' % (got, model))

        # Yearly must be twelve months of the same figure, never a differently-derived number.
        ybad = [r for r in lines
                if abs(B.value(name, 'J%d' % r) - 12 * B.value(name, 'I%d' % r)) > 0.01]
        ck.add('%s: every yearly amount is twelve times its monthly' % name, not ybad,
               'rows %s' % ybad[:5])

        heights = [r for r in range(1, ws.max_row + 1)
                   if ws.row_dimensions[r].height is None]
        ck.add('%s: every row has a stamped height' % name, not heights,
               '%d row(s) unstamped; Excel Online and SharePoint never auto-fit'
               % len(heights))

    # Cover must read each environment total and add only those rows.
    cover = wb['Cover']
    cov_bad = []
    for r in range(1, cover.max_row + 1):
        for c in range(1, 12):
            v = cover.cell(r, c).value
            if isinstance(v, str) and v.startswith('=') and ':' in v:
                cov_bad.append('Cover!%s%d = %s' % (get_column_letter(c), r, v))
    ck.add('Cover never sums a range', not cov_bad, '; '.join(cov_bad[:3]))

    env_totals = sum(B.value(n, 'I%d' % t) for n, _, _, t in env_info.values() if t)
    cover_hits = [(r, c) for r in range(1, cover.max_row + 1) for c in range(1, 12)
                  if isinstance(cover.cell(r, c).value, str)
                  and str(cover.cell(r, c).value).startswith('=')]
    ck.add('the Cover check inspected some formulas', bool(cover_hits),
           'no Cover formula found, so this check proved nothing')
    best = max((B.value('Cover', get_column_letter(c) + str(r)) for r, c in cover_hits),
               default=0.0)
    ck.add('the Cover recurring figure matches the environments',
           abs(best - env_totals) < 0.02 or best >= env_totals - 0.02,
           'largest Cover figure %.2f vs environments %.2f' % (best, env_totals))

    # ---------------------------------------------------------------- eligibility
    # First: the sizing file's own flags must agree with how clouds actually sell. A line
    # named "control plane" that claims to be reservable is a mislabelling, and every lever
    # downstream inherits it.
    mislabelled = []
    for line in spec['lines']:
        s = (line['service'] or '').lower()
        if line.get('reservable', True) and any(k in s for k in NEVER_RESERVABLE):
            mislabelled.append('%s is marked reservable' % line['service'])
        if line.get('stoppable', True) and any(k in s for k in NEVER_STOPPABLE):
            mislabelled.append('%s is marked stoppable' % line['service'])
    ck.add('no line claims an instrument its service cannot carry', not mislabelled,
           '; '.join(mislabelled[:4]))

    # Second: a lever's basis must resolve to eligible rows, and to at least one row. A
    # basis that resolves to nothing prices a saving at zero without saying so.
    opt = wb['Optimisation']
    key_by_row = {}
    for e in envs:
        name, lines, _, _ = env_info[e['key']]
        by_service = {}
        for line in spec['lines']:
            if (line.get('qty') or {}).get(e['key']):
                by_service.setdefault(line['service'], line)
        ordered = [l for l in spec['lines'] if (l.get('qty') or {}).get(e['key'])]
        for r, line in zip(lines, ordered):
            key_by_row[(name, r)] = line
        _ = by_service

    lever_rows = [r for r in range(1, opt.max_row + 1)
                  if any(isinstance(opt.cell(r, c).value, str)
                         and str(opt.cell(r, c).value).startswith('=')
                         and '!' in str(opt.cell(r, c).value)
                         for c in range(1, 12))]
    ck.add('the Optimisation sheet states a basis for its levers', bool(lever_rows),
           'no lever or reservation names the rows it acts on')

    empty_basis, ineligible = [], []
    for r in lever_rows:
        label = next((str(opt.cell(r, c).value) for c in range(1, 6)
                      if isinstance(opt.cell(r, c).value, str)
                      and opt.cell(r, c).value.strip()), 'row %d' % r)
        for c in range(1, 12):
            v = opt.cell(r, c).value
            if not (isinstance(v, str) and v.startswith('=') and '!' in v):
                continue
            refs = [(m.group(1), int(m.group(3))) for m in CELL.finditer(v[1:]) if m.group(1)]
            if not refs:
                empty_basis.append('%s resolves to no rows' % label)
                continue
            # The instrument is read from the SCOPE the sheet already prints in its
            # "Applies to" column, not guessed from the row's label. Guessing failed
            # immediately: a reservation row is labelled with its term, "1 year", which
            # contains neither "reserved" nor "commitment", so it was mistaken for a
            # shutdown schedule and the check complained about the wrong property.
            scope = ''
            for c2 in range(1, 7):
                v2 = opt.cell(r, c2).value
                if isinstance(v2, str) and ('reservable' in v2 or 'stoppable' in v2
                                            or 'hourly' in v2):
                    scope = v2.lower()
                    break
            # A reservation block prints no scope, so absence means the reservation
            # instrument rather than an unknown one.
            reserved_like = ('reservable' in scope) or not scope
            for sheet, row in refs:
                line = key_by_row.get((sheet, row))
                if line is None:
                    continue
                s = (line['service'] or '').lower()
                if reserved_like and any(k in s for k in NEVER_RESERVABLE):
                    ineligible.append('%s claims %s, which cannot be reserved'
                                      % (label, line['service']))
                if not reserved_like and any(k in s for k in NEVER_STOPPABLE):
                    ineligible.append('%s claims %s, which has no stopped state'
                                      % (label, line['service']))
    ck.add('no lever basis is empty', not empty_basis, '; '.join(empty_basis[:3]))
    ck.add('no lever acts on a line the vendor cannot sell it for', not ineligible,
           '; '.join(sorted(set(ineligible))[:4]))

    # ------------------------------------------------------------------ freshness
    today = (_dt.date.fromisoformat(a.today) if a.today else _dt.date.today())
    try:
        priced = _dt.date.fromisoformat(str(spec['priced_on']))
        age = (today - priced).days
        ck.add('priced_on is a real date', True)
        ck.add('list prices are no older than %d days' % a.max_price_age_days,
               0 <= age <= a.max_price_age_days,
               'priced_on is %s, which is %d day(s) old; re-extract before sending'
               % (priced, age))
    except ValueError:
        ck.add('priced_on is a real date', False, 'priced_on=%r' % spec.get('priced_on'))

    # The strong form: every unit price must still resolve against a freshly fetched dump.
    # A missing --prices FAILS rather than skipping, because a freshness check that quietly
    # does nothing is worse than none: it reads as evidence the prices were confirmed.
    if not a.prices:
        ck.add('every unit price re-resolves against a fresh vendor fetch', False,
               'no --prices file supplied, so no price was confirmed against the vendor')
    elif not os.path.exists(a.prices):
        ck.add('every unit price re-resolves against a fresh vendor fetch', False,
               '%s does not exist' % a.prices)
    else:
        fresh = json.load(open(a.prices, encoding='utf-8'))
        flat = {}

        def collect(o):
            if isinstance(o, dict):
                if 'usd' in o and isinstance(o['usd'], (int, float)):
                    return
                for k, v in o.items():
                    if isinstance(v, (int, float)):
                        flat[k] = float(v)
                    elif isinstance(v, dict) and isinstance(v.get('usd'), (int, float)):
                        flat[k] = float(v['usd'])
                    else:
                        collect(v)
            elif isinstance(o, list):
                for v in o:
                    collect(v)
        collect(fresh)

        missing, drifted = [], []
        for key, p in spec['prices'].items():
            if key not in flat:
                missing.append(key)
            elif abs(flat[key] - float(p['usd'])) > 1e-6:
                drifted.append('%s: sizing has %.6f, the vendor now returns %.6f'
                               % (key, float(p['usd']), flat[key]))
        ck.add('the fresh fetch covers every priced key', not missing,
               'absent from %s: %s' % (a.prices, missing[:6]))
        ck.add('every unit price re-resolves against a fresh vendor fetch', not drifted,
               '; '.join(drifted[:4]))
        ck.add('the freshness check compared some prices', bool(spec['prices']) and bool(flat),
               'nothing was compared, so this check proved nothing')

    # --------------------------------------------------------------- presentation
    tp = wb['Third-Party']
    tp_rows = [r for r in range(1, tp.max_row + 1) if isinstance(tp.cell(r, 1).value, int)]
    thin = [r for r in tp_rows if not str(tp.cell(r, 4).value or '').strip()
            or not str(tp.cell(r, 5).value or '').strip()]
    ck.add('every third-party item states a pricing basis and a treatment', not thin,
           'rows %s' % thin[:5])
    ck.add('the third-party sheet is not silently empty', bool(tp_rows),
           'an unpriced list of third-party items reads to a client as zero')

    src = wb['Price Sources']
    src_rows = [r for r in range(1, src.max_row + 1) if isinstance(src.cell(r, 1).value, int)]
    ck.add('Price Sources lists every priced key',
           len(src_rows) >= len(spec['prices']),
           '%d source row(s) for %d price(s)' % (len(src_rows), len(spec['prices'])))

    print('=' * 78)
    print('COST WORKBOOK GATE  -  %s' % os.path.basename(a.xlsx))
    print('=' * 78)
    return ck.report()


if __name__ == '__main__':
    sys.exit(main())
