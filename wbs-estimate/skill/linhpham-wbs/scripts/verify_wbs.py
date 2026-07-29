#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""Gate the WBS workbook before anyone is told it is finished.

    python verify_wbs.py --spec wbs.json --xlsx "WBS_Acme.xlsx"

Exit 0 only when every check passes. Three groups:

  Traceability   every mandatory requirement reaches at least one task, and no task
                 claims a requirement that does not exist
  Structure      whole hours, ascending ids, no duplicates, no orphan rows, deliberate
                 zeros actually zero, roll-ups that cannot double-count
  Rendering      every row has a stamped height and every header fits, because Excel
                 Online and SharePoint never auto-fit and a clipped cell is a defect
                 the author never sees locally

Counting tags is necessary but not sufficient: it proves a reference was mentioned, not
that the work is there. The report prints per-requirement hit counts and the thinnest
coverage so a requirement covered once, in passing, is visible for review.
"""
import argparse
import json
import os
import re
import sys

import openpyxl
from openpyxl.utils import get_column_letter

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
import wbs_schema  # noqa: E402
from xlsx_style import header_fits  # noqa: E402

WIDTHS = {'A': 10.43, 'B': 33.29, 'C': 63.0, 'D': 87.0, 'E': 60.14,
          'F': 12.57, 'G': 15.71, 'H': 15.71, 'I': 10.5, 'J': 9.0, 'K': 12.14}


class Checks:
    def __init__(self):
        self.rows, self.failures = [], []

    def add(self, name, ok, detail=''):
        self.rows.append((name, bool(ok)))
        if not ok:
            self.failures.append(detail or name)
        return ok

    def report(self):
        for name, ok in self.rows:
            print('  [%s] %s' % ('PASS' if ok else 'FAIL', name))
        print()
        if self.failures:
            print('RESULT: %d FAILURE(S)' % len(self.failures))
            for f in self.failures:
                print('  - %s' % f)
            return 1
        print('RESULT: ALL CHECKS PASS (%d checks)' % len(self.rows))
        return 0


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument('--spec', required=True)
    ap.add_argument('--xlsx', required=True)
    ap.add_argument('--fill', action='store_true',
                    help="the workbook is the client's own. Their layout is not this "
                         "skill's, so the structural checks do not apply; the numbers, "
                         'the factor stack and the rendering still do')
    a = ap.parse_args()
    spec = json.load(open(a.spec, encoding='utf-8'))
    ck = Checks()

    # ---------------------------------------------------------------- spec
    print('=' * 78)
    print('1. SPEC')
    print('=' * 78)
    try:
        wbs_schema.validate(spec)
        ck.add('spec is internally consistent', True)
    except wbs_schema.SpecError as e:
        ck.add('spec is internally consistent', False, 'spec invalid:\n%s' % e)
        print(e)

    cols = spec['columns']
    leaves = wbs_schema.leaf_rows(spec)
    print('Leaf tasks : %d' % len(leaves))

    # ---------------------------------------------------------------- traceability
    print()
    print('=' * 78)
    print('2. REQUIREMENT TRACEABILITY')
    print('=' * 78)
    reqs = spec.get('requirements') or []
    if not reqs:
        print('No requirement list in the spec, so coverage cannot be proven.')
        print('Populate `requirements` during ingest: on an authored WBS this is the only')
        print('thing standing between the estimate and a silently missing feature.')
        ck.add('requirement list present', False,
               'spec.requirements is empty; coverage is unproven')
    else:
        used = {}
        for leaf in leaves:
            for ref in leaf.get('refs') or []:
                used.setdefault(ref, []).append(str(leaf['id']))
        known = {r['id'] for r in reqs}
        mandatory = [r for r in reqs if str(r.get('priority', 'M')).upper().startswith('M')]
        missing = [r['id'] for r in mandatory if r['id'] not in used]
        unknown = sorted(set(used) - known)

        print('Requirements     : %d (%d mandatory)' % (len(reqs), len(mandatory)))
        print('Mandatory covered: %d' % (len(mandatory) - len(missing)))
        ck.add('every mandatory requirement reaches a task', not missing,
               'uncovered: %s' % ', '.join(missing[:12]))
        ck.add('no task claims an unknown requirement', not unknown,
               'unknown refs: %s' % ', '.join(unknown[:12]))

        by_ref = {r['id']: r for r in reqs}
        scored = sorted(((sum(hours(by_id(spec, t), cols) for t in used.get(r['id'], [])),
                          len(used.get(r['id'], [])), r['id'],
                          str(r.get('priority', '')), (r.get('text') or '')[:64])
                         for r in mandatory), key=lambda x: x[0])
        print()
        print('Thinnest coverage (hours ascending) - review, do not assume:')
        print('  %-9s %-4s %5s %7s  %s' % ('Ref', 'Pri', 'Tasks', 'Hours', 'Requirement'))
        for h, n, rid, pri, text in scored[:10]:
            print('  %-9s %-4s %5d %7d  %s' % (rid, pri, n, h, text))
        _ = by_ref

    # ---------------------------------------------------------------- workbook
    print()
    print('=' * 78)
    print('3. WORKBOOK')
    print('=' * 78)
    wb = openpyxl.load_workbook(a.xlsx)
    sheet_names = [] if a.fill else [s['name'][:31] for s in spec['sheets']]
    if not a.fill:
        ck.add('sheet order Cover / <line items> / Out of Scope',
               wb.sheetnames == ['Cover'] + sheet_names + ['Out of Scope'],
               'sheets are %s' % wb.sheetnames)

    n_eff = len(cols)
    # The Cover is where a double count actually shows up, and there was no check on it at
    # all. Every figure there must read a single cell on a WBS sheet, or add named cells on
    # the Cover itself. A range is banned outright: once a section row carries its own
    # roll-up, a range over the body spans task, group and module level at once.
    #
    # Scanning every formula column matters. An earlier version of this check looked only at
    # one column, so a range restored into a different column passed silently.
    if not a.fill:
        cover = wb['Cover']
        cover_bad = []
        for r in range(1, cover.max_row + 1):
            for c in range(2, 2 + n_eff + 1):
                v = cover.cell(r, c).value
                if not isinstance(v, str) or not v.startswith('='):
                    continue
                body = v[1:]
                if 'SUM(' in body or ':' in body:
                    cover_bad.append('Cover!%s%d = %s' % (get_column_letter(c), r, v))
                    continue
                for term in body.split('+'):
                    term = term.strip()
                    if not re.fullmatch(r"(?:'[^']+'!)?\$?[A-Z]{1,3}\$?\d+", term):
                        cover_bad.append('Cover!%s%d has the term %r'
                                         % (get_column_letter(c), r, term))
        ck.add('Cover reads single cells and never sums a range', not cover_bad,
               '; '.join(cover_bad[:3]))
        ck.add('the Cover check inspected some formulas',
               any(isinstance(cover.cell(r, c).value, str)
                   and str(cover.cell(r, c).value).startswith('=')
                   for r in range(1, cover.max_row + 1)
                   for c in range(2, 2 + n_eff + 1)),
               'no Cover formula was found, so this check proved nothing')
    first_eff, total_col = 6, 6 + n_eff
    written, zero_written, total_leaf_rows = [], [], 0

    for name in sheet_names:
        ws = wb[name]
        tag = name[:18]
        ck.add('%s: gridlines hidden' % tag, ws.sheet_view.showGridLines is False)

        leaf_rows_, sec_rows, total_row = [], [], None
        for r in range(3, ws.max_row + 1):
            v = ws.cell(r, 1).value
            if ws.cell(r, 2).value == 'TOTAL' and v is None:
                total_row = r
                continue
            t = '' if v is None else str(v)
            if re.fullmatch(r'\d+(\.\d+)?', t) and not ws.cell(r, 3).value:
                sec_rows.append(r)
            elif re.fullmatch(r'\d+(\.\d+)+', t):
                leaf_rows_.append(r)
        total_leaf_rows += len(leaf_rows_)

        # A section row sums its DIRECT children, named cell by cell. The previous rule
        # required these rows to be EMPTY, which is only safe while nobody adds a subtotal;
        # when someone did, every range-based total counted the same hours three times.
        def _direct_children(sec_row):
            sec_id = str(ws.cell(sec_row, 1).value or '').strip()
            depth = sec_id.count('.')
            kids = []
            for rr in range(sec_row + 1, (total_row or ws.max_row) + 1):
                rid = str(ws.cell(rr, 1).value or '').strip()
                if not rid or not rid[0].isdigit():
                    continue
                d = rid.count('.')
                if d <= depth:
                    break
                if d == depth + 1 and rid.startswith(sec_id + '.'):
                    kids.append(rr)
            return kids

        bad_roll = []
        for r in sec_rows:
            kids = _direct_children(r)
            if not kids:
                continue
            for c in range(first_eff, total_col):
                letter = get_column_letter(c)
                want = '=' + '+'.join('%s%d' % (letter, k) for k in kids)
                if ws.cell(r, c).value != want:
                    bad_roll.append('%s!%s%d is %r, expected %r'
                                    % (tag, letter, r, ws.cell(r, c).value, want))
        ck.add('%s: section rows sum their direct children' % tag, not bad_roll,
               '; '.join(bad_roll[:3]))

        # The bug class itself. Once a section row carries a total, any SUM over a range of
        # body rows counts the same hours at every level it spans, so no effort cell may do
        # it. A leaf's =SUM(F<r>:J<r>) is a horizontal sum across one row and is fine.
        vertical = []
        for rr in range(3, ws.max_row + 1):
            for c in range(first_eff, total_col + 1):
                v = ws.cell(rr, c).value
                if not isinstance(v, str) or 'SUM(' not in v:
                    continue
                horizontal = '=SUM(%s%d:%s%d)' % (get_column_letter(first_eff), rr,
                                                  get_column_letter(total_col - 1), rr)
                if v == horizontal:
                    continue
                vertical.append('%s!%s%d = %s' % (tag, get_column_letter(c), rr, v))
        ck.add('%s: no effort total sums a vertical range' % tag, not vertical,
               '; '.join(vertical[:3]))
        ck.add('%s: every leaf total is a formula' % tag,
               all(str(ws.cell(r, total_col).value or '').startswith('=SUM(')
                   for r in leaf_rows_))
        ck.add('%s: TOTAL row present' % tag, total_row is not None)
        if total_row:
            l1 = [r for r in sec_rows
                  if '.' not in str(ws.cell(r, 1).value or '').strip()]
            ck.add('%s: TOTAL sums the level-1 rows' % tag,
                   all(ws.cell(total_row, c).value ==
                       '=' + '+'.join('%s%d' % (get_column_letter(c), x) for x in l1)
                       for c in range(first_eff, total_col)),
                   'TOTAL must add the module rows, not a range that also spans them')

        for r in leaf_rows_:
            for c in range(first_eff, total_col):
                v = ws.cell(r, c).value
                if isinstance(v, (int, float)):
                    written.append((name, r, get_column_letter(c), v))

        no_h = [r for r in range(1, ws.max_row + 1) if ws.row_dimensions[r].height is None]
        ck.add('%s: every row has a stamped height' % tag, not no_h,
               '%s rows without a height: %s (SharePoint will clip them)'
               % (tag, no_h[:6]))
        clipped = header_fits(ws, WIDTHS)
        ck.add('%s: every column header fits' % tag, not clipped,
               '%s: %s' % (tag, '; '.join(clipped[:3])))

        ck.add('%s: column E carries the assumptions' % tag,
               all(str(ws.cell(r, 5).value or '').startswith('Assumption')
                   for r in leaf_rows_),
               '%s has a leaf whose column E does not start with the assumptions' % tag)

    if not a.fill:
        ck.add('leaf rows across sheets match the spec', total_leaf_rows == len(leaves),
               'workbook has %d leaf rows, spec has %d' % (total_leaf_rows, len(leaves)))
    ck.add('all written hours are whole numbers',
           all(float(v) == int(v) for _n, _r, _c, v in written),
           'fractional: %s' % [x for x in written if float(x[3]) != int(x[3])][:5])
    ck.add('no zero written into an effort cell',
           all(v != 0 for _n, _r, _c, v in written),
           'a deliberate zero should be an empty cell, not a 0')
    _ = zero_written

    zcols = spec.get('zero_columns') or []
    if zcols:
        idx = {c: get_column_letter(first_eff + i) for i, c in enumerate(cols)}
        offenders = [x for x in written if x[2] in {idx[c] for c in zcols}]
        ck.add('zeroed columns carry no hours', not offenders,
               'zeroed column has values: %s' % offenders[:4])
    # Read the WORKBOOK, not the spec. The spec still holds the pre-factor numbers because
    # the zeroing happens during the build, so checking it here would pass while the
    # rendered file was wrong. Verify the thing that ships.
    for m in spec.get('zero_modules') or []:
        prefix = str(m) + '.'
        found, nonzero = 0, []
        for name in sheet_names:
            ws = wb[name]
            for r in range(3, ws.max_row + 1):
                rid = ws.cell(r, 1).value
                if rid is None or not ws.cell(r, 3).value:
                    continue
                rid = str(rid)
                if rid != str(m) and not rid.startswith(prefix):
                    continue
                found += 1
                for c in range(first_eff, total_col):
                    if isinstance(ws.cell(r, c).value, (int, float)):
                        nonzero.append('%s!%s' % (name, rid))
                        break
        ck.add('module %s is zeroed in the workbook' % m, found and not nonzero,
               'module %s: %d row(s) rendered, %d still carrying hours %s'
               % (m, found, len(nonzero), nonzero[:4]))
        ck.add('module %s keeps its rows so the scope stays visible' % m, found > 0,
               'module %s has no rows in the workbook at all' % m)

    # risk register
    with_risk = [r for r in leaves if r.get('risk')]
    if with_risk:
        rendered = {}
        for name in sheet_names:
            ws = wb[name]
            for r in range(3, ws.max_row + 1):
                rid = ws.cell(r, 1).value
                if rid is not None and ws.cell(r, 3).value:
                    rendered[str(rid)] = ws.cell(r, 5).value or ''
        missing = [str(r['id']) for r in with_risk
                   if 'Technical risk (' not in rendered.get(str(r['id']), '')]
        ck.add('every registered risk is rendered in column E', not missing,
               'risk block missing for: %s' % ', '.join(missing[:8]))


    if a.fill:
        # Their layout is not this skill's, so nothing structural is asserted. What is
        # asserted is the only thing the fill was for: that every hour in the estimate is
        # present in the file that goes out, at the identifier it belongs to.
        import fill_wbs as _fw
        landed, wrong, hdr_found = 0, [], False
        by_id_spec = {str(r['id']).strip(): r for r in leaves}
        for name in wb.sheetnames:
            ws = wb[name]
            hdr, id_col, found = _fw.find_layout(ws, cols)
            if not hdr or not found:
                continue
            hdr_found = True
            for r in range(hdr + 1, ws.max_row + 1):
                rid = str(ws.cell(r, id_col).value or '').strip()
                leaf = by_id_spec.get(rid)
                if not leaf:
                    continue
                for key, c in found.items():
                    want_v = leaf.get(key) or 0
                    got_v = ws.cell(r, c).value or 0
                    if not isinstance(got_v, (int, float)):
                        got_v = 0
                    if int(got_v) != int(want_v):
                        wrong.append('%s!%s row %d: %s is %s, the estimate says %s'
                                     % (name, rid, r, key, got_v, want_v))
                landed += 1
        ck.add('the effort columns were found in the delivered workbook', hdr_found,
               'no sheet had a header row naming the effort columns, so nothing was checked')
        ck.add('every estimated task appears in the delivered workbook',
               landed == len(leaves),
               '%d of %d task(s) found by identifier' % (landed, len(leaves)))
        ck.add('every hour in the delivered workbook matches the estimate', not wrong,
               '; '.join(wrong[:4]))
        heights = [n for n in wb.sheetnames
                   for r in range(1, wb[n].max_row + 1)
                   if wb[n].row_dimensions[r].height is None]
        ck.add('every row has a stamped height', not heights,
               '%d row(s) unstamped; Excel Online and SharePoint never auto-fit' % len(heights))

    # ------------------------------------------------------- the factor stack has two sides
    # This is the single most expensive rule in the reference, and it was documented and not
    # enforced. An estimate that applies the AI discount and nothing else silently drops the
    # integration buffer, the no-sandbox uplift, the legacy-protocol uplift and the rule that
    # a row bundling N integrations costs N times one. On the bid this was ported from it lost
    # 218 hours, and it was only found because someone asked whether the rules had been
    # applied at all. Nothing in the workbook looks wrong when it happens.
    factors = spec.get('factors') or []
    upward = [f for f in factors if (f.get('final') or 0) > (f.get('base') or 0)]
    downward = [f for f in factors if (f.get('final') or 0) < (f.get('base') or 0)]
    unc = (spec.get('uncertainty') or {}).get('factor')
    has_upward = bool(upward) or (unc and unc > 1.0)

    # Only demand an uplift where the estimate itself says one is owed. A project with no
    # external interface genuinely needs none, and a blanket rule would cry wolf on it.
    OWES = re.compile(
        r'\b(two|three|four|five|\d+)\s+\w*\s*'
        r'(integration|integrations|interface|interfaces|gateway|gateways|bank|banks|'
        r'authority|authorities|rail|rails|partner|partners|provider|providers|'
        r'product|products|system|systems)\b', re.I)
    owed = []
    for leaf in leaves:
        text = ' '.join(str(leaf.get(k) or '') for k in ('assum', 'desc', 'feature'))
        m = OWES.search(text)
        if m and not any(str(f['id']) == str(leaf['id']) for f in factors):
            owed.append('%s (%s)' % (leaf['id'], m.group(0).strip()))

    # An empty factor table must NOT pass this vacuously. That is precisely the state the bid
    # was in when it was 218 hours short: the AI discount had been folded into the base
    # numbers as the estimating phase intends, so the factor table was empty and there was
    # nothing to look one-sided. The integration buffer is owed on every task that talks to
    # something outside the estate, so the presence of such a task is what makes the demand.
    INTEGRATES = re.compile(
        r'\b(integrat\w*|gateway|webhook|connector|third.part\w*|external|api of|sso|'
        r'oauth|soap|sftp|erp|hrms|payment|acquirer|authority|bank|partner system)\b', re.I)
    integrating = [str(l['id']) for l in leaves
                   if INTEGRATES.search(' '.join(str(l.get(k) or '')
                                                 for k in ('assum', 'desc', 'feature')))]
    ck.add('the factor stack moves in both directions, not only down',
           has_upward or not integrating,
           'the factor table has %d downward and %d upward entries while %d task(s) integrate '
           'with something outside the estate (%s). Rule 5 owes those a 10-15%% buffer, and '
           'the uplift checklist in reference/estimation_rules.md section 4 owes more: the '
           'no-sandbox multiplier, legacy protocols, and any row that bundles several '
           'integrations. An empty factor table is not evidence none were needed'
           % (len(downward), len(upward), len(integrating), ', '.join(integrating[:6])))
    ck.add('every row bundling several integrations carries an explicit factor', not owed,
           'these name a COUNT of external things but appear in no factor entry, so each is '
           'priced as one: %s' % '; '.join(owed[:6]))

    check_against_reference(spec, leaves, cols, ck)

    # ---------------------------------------------------------------- ratios
    print()
    print('=' * 78)
    print('4. SANITY (report, not pass/fail: an outlier is a question to answer)')
    print('=' * 78)
    grand = sum(r.get(c) or 0 for r in leaves for c in cols)
    cells = sum(1 for r in leaves for c in cols if r.get(c))
    if grand and cells:
        print('  Total                       : %d h' % grand)
        print('  Average per leaf task       : %.1f h   (target 4-7)' % (grand / len(leaves)))
        print('  Average per populated cell  : %.1f h   (target 4-7)' % (grand / cells))
        for c in cols:
            s = sum(r.get(c) or 0 for r in leaves)
            print('  %-27s %d h  %.1f%%' % (c, s, 100.0 * s / grand))
        be = sum(r.get('be') or 0 for r in leaves)
        client = sum((r.get('fe') or 0) + (r.get('mob') or 0) for r in leaves)
        if be + client:
            print('  Back-end vs client          : %.1f%% / %.1f%%'
                  % (100.0 * be / (be + client), 100.0 * client / (be + client)))
        # Two shares the reference calls out by name, so they are printed rather than left
        # for the reader to compute from the module list.
        infra = sum(r.get(c) or 0 for r in leaves for c in cols
                    if wbs_schema.module_of(r) == '1')
        print('  Infrastructure share        : %.1f%%  (target 8-12; below 5 usually means '
              'setup is missing, above 15 usually means it is over-built)'
              % (100.0 * infra / grand))
        nfr_mod = [m for m in sorted({wbs_schema.module_of(r) for r in leaves})][-1:]
        if nfr_mod:
            nfr = sum(r.get(c) or 0 for r in leaves for c in cols
                      if wbs_schema.module_of(r) == nfr_mod[0])
            zeroed = nfr_mod[0] in [str(m) for m in spec.get('zero_modules', [])]
            print('  Last module (%s) share       : %.1f%%%s'
                  % (nfr_mod[0], 100.0 * nfr / grand,
                     '  (deliberately zeroed; the rows are kept)' if zeroed
                     else '  (target 10-15 when it holds the non-functional work)'))
        if 'mob' in cols and 'fe' in cols:
            fe = sum(r.get('fe') or 0 for r in leaves)
            mob = sum(r.get('mob') or 0 for r in leaves)
            if fe and mob:
                print('  Mobile vs front-end web     : %.2fx  (expect 1.2-1.5x for the same '
                      'feature set; below 1.0 usually means the mobile-only work was missed)'
                      % (mob / fe))
        print()
        print('  Report both averages. A leaf spanning several disciplines shows a high')
        print('  average per row and a normal average per cell; quoting only the first')
        print('  invites a correction that is not needed.')

    print()
    print('=' * 78)
    return ck.report()


def reference_ranges():
    """Read the section 3 hour ranges out of the reference file itself.

    Parsed rather than duplicated here, so the table an author reads and the range a check
    enforces cannot drift apart. Returns [(words, be_lo, be_hi, fe_lo, fe_hi)].
    """
    path = os.path.join(os.path.dirname(os.path.abspath(__file__)),
                        '..', 'reference', 'estimation_rules.md')
    if not os.path.exists(path):
        return []
    text = open(path, encoding='utf-8').read()
    try:
        seg = text[text.index('## 3. Effort reference'):text.index('## 4. Factors')]
    except ValueError:
        return []
    out = []
    for line in seg.splitlines():
        if not line.startswith('|') or '---' in line:
            continue
        cells = [c.strip() for c in line.strip('|').split('|')]
        if len(cells) < 2:
            continue

        def span(v):
            m = re.match(r'^(\d+)\s*-\s*(\d+)', v)
            if m:
                return int(m.group(1)), int(m.group(2))
            if re.fullmatch(r'\d+', v):
                return int(v), int(v)
            return None

        be = span(cells[1]) if len(cells) > 1 else None
        fe = span(cells[2]) if len(cells) > 2 else None
        if be is None and fe is None:
            continue
        words = {w for w in re.findall(r'[a-z]{4,}', cells[0].lower())
                 if w not in ('task', 'with', 'from', 'each', 'that', 'this', 'when')}
        if words:
            out.append((words, be, fe))
    return out


def check_against_reference(spec, leaves, cols, ck):
    """Flag a task priced under the range its own shape carries in the reference.

    Only the low side fails. Being under the reference range is how a bid is won and then
    lost, and it is invisible once the workbook totals up. Being over is a question for the
    review, not a defect, because a leaf legitimately bundles several reference tasks.

    The reference ranges are pre-discount, so the comparison is made against the range after
    the declared AI factor. Comparing raw would flag most of a well-formed estimate.
    """
    ranges = reference_ranges()
    if not ranges:
        ck.add('the reference ranges could be read', False,
               'reference/estimation_rules.md section 3 did not parse, so no task was '
               'compared against its own shape')
        return
    ai = (spec.get('ai_factor') or {})
    scale = ai.get('blended') if ai.get('where') != 'none' else 1.0
    scale = scale or 1.0

    under, matched = [], 0
    for leaf in leaves:
        text = ' '.join(str(leaf.get(k) or '') for k in ('feature', 'group')).lower()
        tw = {w for w in re.findall(r'[a-z]{4,}', text)}
        if not tw:
            continue
        best, hit = 0.0, None
        for words, be, fe in ranges:
            score = len(words & tw) / len(words)
            if score > best:
                best, hit = score, (words, be, fe)
        if best < 0.6 or hit is None:
            continue
        matched += 1
        _words, be, fe = hit
        for col, rng in (('be', be), ('fe', fe)):
            if col not in cols or not rng:
                continue
            floor = rng[0] * scale
            got = leaf.get(col) or 0
            if got and got < floor * 0.75:
                under.append('%s %s=%d, the reference gives %d-%d for "%s" (x%.2f = %.1f)'
                             % (leaf['id'], col, got, rng[0], rng[1],
                                ' '.join(sorted(_words))[:28], scale, floor))
    ck.add('no task is priced under the reference range for its shape', not under,
           '; '.join(under[:4]))
    # Deliberately a report, not a gate. Task naming legitimately varies between clients
    # and domains, so demanding a vocabulary match would fire on most well-formed specs,
    # and a check that cries wolf is worse than no check because it trains the reader to
    # skip the output. The hard gate stays on the one direction that costs money.
    print('  Reference-range comparison   : %d of %d task(s) matched a section 3 shape'
          % (matched, len(leaves)))
    if not matched:
        print('  None matched, so this run compared nothing. Naming the work the way')
        print('  section 3 names it makes the comparison possible.')


def by_id(spec, task_id):
    for r in spec['rows']:
        if str(r.get('id')) == str(task_id):
            return r
    return {}


def hours(row, cols):
    return sum(row.get(c) or 0 for c in cols)


if __name__ == '__main__':
    sys.exit(main())
