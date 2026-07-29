#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""Prove verify_cost.py catches the defects it claims to.

    python calibrate_cost.py --sizing sizing.json --xlsx cost.xlsx --prices prices.json

A check that has only ever seen a clean workbook is not a check. Each fault below is a real
way a cost estimate goes wrong, and several of them shipped on a live bid before the
corresponding check existed:

  * a total that misses a layer, which reads as a cheaper estate
  * a monthly amount pasted as a literal, so a corrected quantity changes nothing
  * a saving claimed on a line the vendor does not sell that instrument for
  * a list price carried forward from weeks ago under today's date
  * a price that has since moved at the vendor
  * a freshness check invoked with no vendor file, which must FAIL rather than skip

Every mutation is applied to a copy. Nothing here touches the real workbook.
"""
import argparse
import copy
import json
import os
import shutil
import subprocess
import sys
import tempfile

from openpyxl import load_workbook

HERE = os.path.dirname(os.path.abspath(__file__))
VERIFY = os.path.join(HERE, 'verify_cost.py')


def run(sizing, xlsx, prices, today):
    cmd = [sys.executable, VERIFY, '--sizing', sizing, '--xlsx', xlsx, '--today', today]
    if prices:
        cmd += ['--prices', prices]
    p = subprocess.run(cmd, capture_output=True, text=True)
    fails = [ln.strip()[7:] for ln in p.stdout.splitlines() if ln.strip().startswith('[FAIL]')]
    crashed = 'Traceback' in (p.stderr or '')
    return p.returncode, fails, crashed


def first_line_row(ws):
    for r in range(1, ws.max_row + 1):
        if str(ws.cell(r, 1).value or '').strip() == '#':
            return r + 1
    return None


def total_row(ws):
    for r in range(1, ws.max_row + 1):
        if str(ws.cell(r, 2).value or '').strip().upper().startswith('TOTAL'):
            return r
    return None


# each fault: (label, mutate_workbook_or_None, mutate_sizing_or_None, drop_prices)
def env_total_drops_a_layer(wb, name):
    ws = wb[name]
    tr = total_row(ws)
    cur = ws.cell(tr, 9).value            # =SUM(I<first>:I<last>)
    inner = cur[cur.index('(') + 1:cur.index(')')]
    a, b = inner.split(':')
    ws.cell(tr, 9, '=SUM(%s:%s%d)' % (a, b[0], int(b[1:]) - 1))


FAULTS = [
    ('environment TOTAL drops a layer',
     lambda wb, n: env_total_drops_a_layer(wb, n), None, False),
    ('a monthly amount pasted as a literal',
     lambda wb, n: wb[n].cell(first_line_row(wb[n]), 9, 123.45), None, False),
    ('a yearly amount that is not twelve months',
     lambda wb, n: wb[n].cell(first_line_row(wb[n]), 10, '=I%d*11' % first_line_row(wb[n])),
     None, False),
    ('a row left with no stamped height',
     lambda wb, n: setattr(wb[n].row_dimensions[first_line_row(wb[n])], 'height', None),
     None, False),
    ('a lever claims a line with no stopped state',
     None, lambda s: s['levers'][0].update({'scope': 'non-prod hourly'}), False),
    ('a line mislabelled as reservable',
     None, lambda s: [l.update({'reservable': True}) for l in s['lines']
                      if 'control plane' in l['service'].lower()], False),
    ('list prices carried forward from weeks ago',
     None, lambda s: s.update({'priced_on': '2026-05-01'}), False),
    ('a unit price that has since moved at the vendor',
     None, lambda s: s['prices'][sorted(s['prices'])[0]].update({'usd': 9.99}), False),
    ('the freshness check invoked with no vendor file',
     None, None, True),
]


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument('--sizing', required=True)
    ap.add_argument('--xlsx', required=True)
    ap.add_argument('--prices', required=True)
    ap.add_argument('--today', default='2026-07-29')
    a = ap.parse_args()

    rc, fails, crashed = run(a.sizing, a.xlsx, a.prices, a.today)
    print('clean workbook: %s' % ('PASS' if rc == 0 and not crashed else 'FAIL (unexpected)'))
    if rc != 0:
        for f in fails:
            print('   ' + f)
        return 1

    base_spec = json.load(open(a.sizing, encoding='utf-8'))
    env0 = base_spec['environments'][0]['name'][:31]
    tmp = tempfile.mkdtemp(prefix='cost_cal_')
    caught = 0
    try:
        for label, mut_wb, mut_spec, drop_prices in FAULTS:
            sz = os.path.join(tmp, 'sizing.json')
            xl = os.path.join(tmp, 'cost.xlsx')
            spec = copy.deepcopy(base_spec)
            if mut_spec:
                mut_spec(spec)
            json.dump(spec, open(sz, 'w', encoding='utf-8'), indent=1)

            if mut_spec and not mut_wb:
                # rebuild so the workbook reflects the mutated spec
                subprocess.run([sys.executable, os.path.join(HERE, 'build_cost.py'),
                                '--sizing', sz, '--out', xl],
                               capture_output=True, text=True)
            else:
                shutil.copy(a.xlsx, xl)
            if mut_wb:
                wb = load_workbook(xl)
                mut_wb(wb, env0)
                wb.save(xl)

            rc, fails, crashed = run(sz, xl, None if drop_prices else a.prices, a.today)
            ok = rc != 0 and not crashed
            caught += ok
            state = 'CAUGHT' if ok else ('*** CRASHED, not caught ***' if crashed
                                         else '*** MISSED ***')
            print('  %-46s %s' % (label, state))
            if ok and fails:
                print('        by: %s' % fails[0])
    finally:
        shutil.rmtree(tmp, ignore_errors=True)

    print('\n%d of %d fault(s) caught' % (caught, len(FAULTS)))
    return 0 if caught == len(FAULTS) else 1


if __name__ == '__main__':
    sys.exit(main())
