#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""Render a WBS workbook from `wbs.json`.

    python build_wbs.py --spec wbs.json --out "WBS_Acme.xlsx"

The estimate lives in the spec; this file only draws it. Two consequences worth
knowing: a re-run of the same spec produces the same workbook, and changing a number
never means editing code.

The factor layer is applied here rather than baked into the spec's numbers, so the
build prints a base -> final -> rule table. An estimate whose reasoning is invisible
cannot be reviewed, and the most expensive mistake in the reference log was an estimate
that applied only the downward factor because nothing showed the upward ones missing.

Sheet layout follows the reference workbooks: one sheet per commercial line item, plus
a Cover that rolls up and an Out of Scope sheet that draws the boundary.
"""
import argparse
import json
import os
import sys

from openpyxl import Workbook
from openpyxl.utils import get_column_letter

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))

from xlsx_style import (  # noqa: E402
    A_CENTER, A_LEFT_CENTER, A_LEFT_TOP, B_LTB, B_NONE, B_RTB, B_TB, C_EFFORT, C_HEADER,
    C_HEADER_TOT, C_L1, C_L2, C_LEAF, C_OOS_HDR, C_TOTAL, FONT, autofit_row_heights,
    banner, notes, style,
)
from openpyxl.styles import Font  # noqa: E402
import wbs_schema  # noqa: E402

WIDTHS = {'A': 10.43, 'B': 33.29, 'C': 63.0, 'D': 87.0, 'E': 60.14,
          'F': 12.57, 'G': 15.71, 'H': 15.71, 'I': 10.5, 'J': 9.0, 'K': 12.14}
COVER_WIDTHS = {'A': 48, 'B': 13, 'C': 15, 'D': 15, 'E': 11, 'F': 9, 'G': 13}
OOS_WIDTHS = {'A': 6.57, 'B': 60.0, 'C': 70.0}
DEFAULT_LABELS = {'ui': 'UI/UX Design', 'be': 'Back-end Development',
                  'fe': 'Front-end Development', 'mob': 'Mobile', 'ai': 'AI'}


SITUATION_FACTOR = {
    'existing client': (0.90, 0.90),
    'follow-on': (0.90, 0.90),
    'new client': (0.85, 0.90),
    'many vendors bidding': (0.85, 0.90),
    'sole bidder': (0.95, 1.00),
    'referral': (0.95, 1.00),
    'exclusive': (0.95, 1.00),
    'complex, few vendors capable': (0.95, 1.00),
}


# ------------------------------------------------------------------ factor layer
def apply_factors(spec):
    """Explicit uplifts, then the selective uncertainty factor, then deliberate zeros."""
    cols = spec['columns']
    by_id = {str(r['id']): r for r in spec['rows']}
    log = []

    for f in spec.get('factors', []):
        row = by_id[str(f['id'])]
        row[f['col']] = f['final']
        log.append(('adjust', f['id'], f['col'], f['base'], f['final'], f.get('rule', ''),
                    f.get('note', '')))

    unc = spec.get('uncertainty') or {}
    factor = unc.get('factor')
    if factor:
        adjusted = {(str(f['id']), f['col']) for f in spec.get('factors', [])}
        for scope in unc.get('scopes', []):
            prefix = scope['prefix'] + '.'
            before = after = 0
            for row in wbs_schema.leaf_rows(spec):
                if not str(row['id']).startswith(prefix):
                    continue
                for c in cols:
                    v = row.get(c) or 0
                    if not v:
                        continue
                    before += v
                    if (str(row['id']), c) in adjusted:
                        after += v          # a stronger rule already moved it
                        continue
                    row[c] = max(1, int(v * factor + 0.5))
                    after += row[c]
            if before:
                log.append(('uncertainty', scope['prefix'], '', before, after,
                            'x%.2f' % factor, scope.get('reason', '')))

    for c in spec.get('zero_columns', []):
        out = sum(r.get(c) or 0 for r in wbs_schema.leaf_rows(spec))
        n = sum(1 for r in wbs_schema.leaf_rows(spec) if r.get(c))
        for r in wbs_schema.leaf_rows(spec):
            r[c] = 0
        log.append(('zero', 'column ' + c, '', out, 0, 'absorbed', '%d row(s)' % n))
    for m in spec.get('zero_modules', []):
        rows_ = [r for r in wbs_schema.leaf_rows(spec) if wbs_schema.module_of(r) == str(m)]
        out = sum(r.get(c) or 0 for r in rows_ for c in cols)
        for r in rows_:
            for c in cols:
                r[c] = 0
        log.append(('zero', 'module ' + str(m), '', out, 0, 'absorbed',
                    '%d row(s), kept so the scope stays visible' % len(rows_)))

    # Section 1 of the rules puts the competitive factor INSIDE the standard formula:
    # base x AI x competitive is the number the client is given. Leaving it at 1.0 by
    # default is how a bid is lost to a cheaper vendor, so when the spec does not set it
    # the situation table decides, and the floors below stop it going too far.
    comp = spec.get('competitive')
    if comp is None:
        sit = str(spec.get('bid_situation') or '').lower()
        for key, (lo, hi) in SITUATION_FACTOR.items():
            if key in sit:
                comp = hi          # the cautious end of the range the table gives
                break
        comp = comp or 1.0
    # Write it back, or the factor log and the competitive position report different
    # numbers in the same run and neither is obviously the wrong one.
    spec['competitive'] = comp
    if comp and comp != 1.0:
        before = sum(r.get(c) or 0 for r in wbs_schema.leaf_rows(spec) for c in cols)
        for r in wbs_schema.leaf_rows(spec):
            for c in cols:
                if r.get(c):
                    r[c] = max(1, int(r[c] * comp + 0.5))
        after = sum(r.get(c) or 0 for r in wbs_schema.leaf_rows(spec) for c in cols)
        # A percentage applied per task, to whole hours, mostly does not survive the
        # rounding. Half-up rounding leaves 1 to 5 hours untouched at x0.90, and a
        # well-formed WBS has most tasks in exactly that band, so the intended discount
        # can arrive as almost nothing. Recording what was actually realised is the
        # difference between giving a discount and believing you gave one.
        realised = (1 - after / before) if before else 0.0
        spec['_competitive_realised'] = realised
        spec['_competitive_before'] = before
        log.append(('competitive', 'all', '', before, after, 'x%.2f' % comp,
                    'intended -%.0f%%, realised -%.1f%%  (%s)'
                    % ((1 - comp) * 100, realised * 100,
                       spec.get('bid_situation') or 'situation not stated')))
    apply_floors(spec, log)
    return log


def print_factor_log(log, comp):
    if not log:
        print('Factor layer: nothing applied.')
        return
    print('=' * 96)
    print('FACTOR LAYER  (base -> final, and the rule that moved it)')
    print('=' * 96)
    print('%-12s %-16s %-5s %8s %8s  %-12s %s'
          % ('kind', 'what', 'col', 'base', 'final', 'rule', 'note'))
    for kind, what, col, base, final, rule, note in log:
        print('%-12s %-16s %-5s %8s %8s  %-12s %s'
              % (kind, str(what)[:16], col, base, final, rule, str(note)[:44]))
    print()
    print('Competitive factor: x%.2f%s'
          % (comp, '' if comp != 1.0 else '  (hours not discounted)'))
    print()



# Section 1 of the rules: the floors are what protect the margin, not refusing the discount.
# Below these a task has no room for its own testing, so the saving is borrowed from quality.
FLOOR_RULES = (
    (('auth', 'login', 'sign-in', 'sso', 'oauth', 'mfa', 'otp', 'password'), 6,
     'security work is not cut'),
    (('integrat', 'gateway', 'webhook', 'connector', 'api of', 'third-part', 'soap',
      'sftp', 'erp', 'hrms'), 6, 'retry, error handling and rate limits'),
    (('infrastructur', 'landing zone', 'pipeline', 'cluster', 'network', 'terraform',
      'iac', 'environment'), 4, 'wrong infrastructure takes the project down'),
    ((), 3, 'under three hours leaves no room to test'),
)


def floor_for(row):
    text = ' '.join(str(row.get(k) or '') for k in ('feature', 'group', 'desc')).lower()
    for words, floor, why in FLOOR_RULES:
        if not words:
            return floor, why
        if any(w in text for w in words):
            return floor, why
    return 3, 'under three hours leaves no room to test'


def apply_floors(spec, log):
    """Raise any leaf the competitive factor pushed under its floor.

    The competitive factor is applied to win the work; the floors are applied so winning it
    is still worth having. They are not in tension: one moves the price, the other refuses to
    move the part of the estimate that pays for testing the work.
    """
    cols = spec['columns']
    zero = set(spec.get('zero_columns') or [])
    zmods = {str(m) for m in (spec.get('zero_modules') or [])}
    for row in wbs_schema.leaf_rows(spec):
        if wbs_schema.module_of(row) in zmods:
            continue
        live = [c for c in cols if c not in zero and (row.get(c) or 0)]
        total = sum(row.get(c) or 0 for c in live)
        if not total:
            continue
        floor, why = floor_for(row)
        if total >= floor:
            continue
        biggest = max(live, key=lambda c: row.get(c) or 0)
        row[biggest] += floor - total
        log.append(('floor', row['id'], biggest, total, floor, '>=%dh' % floor, why))


# ------------------------------------------------------- competitive position


def competitive_position(spec, total):
    """What makes this number competitive, in figures rather than adjectives.

    A percentage taken off the hours is the weakest lever available and the only one that
    cannot be undone later. The strong levers are already in the estimate and normally go
    unstated: the discount the delivery model has already granted, and the work quoted once
    instead of once per product. Both are computed here so they can be quoted in the bid
    instead of reconstructed by hand afterwards.

    Returns (client_facing_lines, internal_lines).
    """
    client, internal = [], []
    cols = spec['columns']

    ai = spec.get('ai_factor') or {}
    blended = ai.get('blended')
    if ai.get('where') != 'none' and blended:
        pre = total / blended
        granted = int(round(pre - total))
        internal.append(('Delivery-model discount already granted',
                         '%d h  (effective x%.2f, %s). Applied %s, so it must not be taken '
                         'again' % (granted, blended, ai.get('note', 'per task type'),
                                    'inside the base numbers'
                                    if ai.get('where') == 'in_base' else 'as factor entries')))

    reuse = spec.get('reuse') or {}
    shared = [n for n in (reuse.get('shared_sheets') or [])]
    consumers = reuse.get('consumers') or 0
    if shared and consumers > 1:
        sheet_mods = {}
        for sh in spec['sheets']:
            sheet_mods[sh['name']] = {str(m) for m in sh['modules']}
        shared_mods = set()
        for n in shared:
            shared_mods |= sheet_mods.get(n, set())
        shared_h = sum(r.get(c) or 0 for r in wbs_schema.leaf_rows(spec) for c in cols
                       if wbs_schema.module_of(r) in shared_mods)
        saved = shared_h * (consumers - 1)
        naive = total + saved
        client.append(('Shared platform quoted once, not %d times' % consumers,
                       '%d h of shared work serves %d products, so %d h are not charged'
                       % (shared_h, consumers, saved)))
        if naive:
            client.append(('Against a build with no reuse',
                           '%d h. This estimate is %.0f%% of it' % (naive, 100.0 * total / naive)))

    comp = spec.get('competitive', 1.0)
    sit = str(spec.get('bid_situation') or '').strip()
    lo = hi = None
    for key, (a, b) in SITUATION_FACTOR.items():
        if key in sit.lower():
            lo, hi = a, b
            break
    line = 'x%.2f' % comp
    if lo is not None:
        line += '. The situation "%s" suggests x%.2f to x%.2f' % (sit, lo, hi)
        if comp > hi:
            line += '. HOLDING ABOVE THE RANGE: say why in the report, or take the discount'
    internal.append(('Competitive factor', line))
    realised = spec.get('_competitive_realised')
    if realised is not None and comp < 1.0:
        intended = 1 - comp
        internal.append(('Discount actually realised',
                         '-%.1f%% of %d h, against -%.0f%% intended'
                         % (realised * 100, spec.get('_competitive_before') or 0,
                            intended * 100)))
        if realised < intended * 0.5:
            internal.append(('WHY THE DISCOUNT DID NOT LAND',
                             'whole-hour rounding absorbed it: at this factor a task '
                             'of 1 to 5 hours does not move, and most tasks are in '
                             'that band. A percentage cannot be expressed per task on '
                             'a fine-grained WBS. Take it as a stated discount on the '
                             'total, on the rate, or by moving scope into phase 1, and '
                             'do not report a reduction the hours do not contain'))
    return client, internal


def phase1_total(spec):
    """Hours in the cheaper option. A subset of the same ids, so the two options can never
    disagree about what a task costs."""
    p1 = spec.get('phase1') or {}
    inc = [str(x) for x in (p1.get('include') or [])]
    if not inc:
        return None, None
    cols = spec['columns']
    rows = [r for r in wbs_schema.leaf_rows(spec)
            if any(str(r['id']) == p or str(r['id']).startswith(p + '.') for p in inc)]
    return sum(r.get(c) or 0 for r in rows for c in cols), len(rows)


def print_competitive_position(spec, total):
    client, internal = competitive_position(spec, total)
    if not (client or internal):
        return
    print('=' * 96)
    print('COMPETITIVE POSITION  (the figures that argue the price, not adjectives)')
    print('=' * 96)
    for label, value in client:
        print('  [bid]      %-44s %s' % (label, value))
    for label, value in internal:
        print('  [internal] %-44s %s' % (label, value))
    p1, n1 = phase1_total(spec)
    if p1:
        # Only printed when the spec actually defines one. Offering a cut-down option is
        # a commercial choice, so the build states it when asked and stays quiet when not.
        p1o = spec['phase1']
        print()
        print('    Full scope   %6d h   %d task(s)'
              % (total, len(wbs_schema.leaf_rows(spec))))
        print('    %-12s %6d h   %d task(s)   %.0f%% of full'
              % (p1o.get('name', 'phase 1')[:12], p1, n1, 100.0 * p1 / total))
        print('    gives up: %s' % p1o.get('note', ''))
    print()
    print('  [bid] lines may be quoted to the client. [internal] lines are grounds for')
    print('  defending the number and are deliberately kept out of the workbook: a delivery')
    print('  discount a client can see is a discount they will ask for twice.')
    print()


# ------------------------------------------------------------------ WBS sheet
def build_sheet(ws, spec, modules):
    cols = spec['columns']
    labels = spec.get('column_labels') or [DEFAULT_LABELS.get(c, c.upper()) for c in cols]
    n_eff = len(cols)
    first_eff = 6
    total_col = first_eff + n_eff              # K when there are five columns
    last_col = total_col

    ws.sheet_view.showGridLines = False
    for letter, width in WIDTHS.items():
        if get_column_letter(last_col) >= letter or letter in 'ABCDE':
            ws.column_dimensions[letter].width = width
    ws.column_dimensions[get_column_letter(total_col)].width = 12.14

    for col, title in enumerate(['ID', 'Module', 'Feature/Screen', 'Description',
                                 'Assumptions, Constraints, Tech Solutions'], start=1):
        ws.cell(1, col, title)
        for row in (1, 2):
            style(ws.cell(row, col), rgb=C_HEADER, bold=True, white=True)
        ws.merge_cells(start_row=1, start_column=col, end_row=2, end_column=col)

    ws.cell(1, first_eff, 'EFFORT (man-hour)')
    for col in range(first_eff, first_eff + n_eff):
        style(ws.cell(1, col), rgb=C_EFFORT, bold=True, white=True)
    ws.merge_cells(start_row=1, start_column=first_eff,
                   end_row=1, end_column=first_eff + n_eff - 1)
    for i, label in enumerate(labels):
        ws.cell(2, first_eff + i, label)
        style(ws.cell(2, first_eff + i), rgb=C_EFFORT, bold=True, white=True)

    ws.cell(1, total_col, 'Total Effort')
    for row in (1, 2):
        style(ws.cell(row, total_col), rgb=C_HEADER_TOT, bold=True, white=True)
    ws.merge_cells(start_row=1, start_column=total_col, end_row=2, end_column=total_col)

    eff_a, eff_z = get_column_letter(first_eff), get_column_letter(first_eff + n_eff - 1)
    r = 3
    spans, groups, group_open = [], [], None
    leaves = []

    def section(item, rgb):
        ws.cell(r, 1, item['id'])
        style(ws.cell(r, 1), rgb=rgb)
        ws.cell(r, 2, item['title'])
        style(ws.cell(r, 2), rgb=rgb, bold=True, align=A_LEFT_CENTER)
        style(ws.cell(r, 3), bd=B_TB)
        style(ws.cell(r, 4), bd=B_TB)
        style(ws.cell(r, 5), bd=B_RTB)
        for c in range(first_eff, last_col + 1):
            style(ws.cell(r, c), rgb=rgb)
        ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=5)

    meta = None
    for item in spec['rows']:
        if int(str(item['id']).split('.')[0]) not in modules:
            continue
        kind = item['kind']
        if kind in ('L1', 'L2'):
            if group_open and group_open[1] > group_open[0]:
                groups.append(tuple(group_open))
            group_open = None
        if kind == 'L1':
            if meta:
                spans.append(meta + (r - 1,))
            meta = (item['id'], item['title'], r)
            section(item, C_L1)
            r += 1
        elif kind == 'L2':
            section(item, C_L2)
            r += 1
        else:
            leaves.append(item)
            ws.cell(r, 1, item['id'])
            style(ws.cell(r, 1), rgb=C_LEAF)
            if item.get('group'):
                if group_open and group_open[1] > group_open[0]:
                    groups.append(tuple(group_open))
                group_open = [r, r]
                ws.cell(r, 2, item['group'])
            elif group_open:
                group_open[1] = r
            style(ws.cell(r, 2), rgb=C_LEAF, align=A_LEFT_TOP)
            ws.cell(r, 3, item['feature'])
            style(ws.cell(r, 3), rgb=C_LEAF, align=A_LEFT_TOP)
            ws.cell(r, 4, item.get('desc', ''))
            style(ws.cell(r, 4), rgb=C_LEAF, align=A_LEFT_TOP)
            ws.cell(r, 5, column_e(item))
            style(ws.cell(r, 5), rgb=C_LEAF, align=A_LEFT_TOP, bd=B_LTB)
            for i, c in enumerate(cols):
                v = item.get(c) or 0
                style(ws.cell(r, first_eff + i, v if v else None), rgb=C_LEAF)
            ws.cell(r, total_col, '=SUM(%s%d:%s%d)' % (eff_a, r, eff_z, r))
            style(ws.cell(r, total_col), rgb=C_LEAF)
            r += 1
    if group_open and group_open[1] > group_open[0]:
        groups.append(tuple(group_open))
    if meta:
        spans.append(meta + (r - 1,))
    for a, b in groups:
        ws.merge_cells(start_row=a, start_column=2, end_row=b, end_column=2)

    last = r - 1
    l1_rows = apply_rollup(ws, first_eff, n_eff, total_col, last)
    style(ws.cell(r, 1), rgb=C_TOTAL)
    ws.cell(r, 2, 'TOTAL')
    style(ws.cell(r, 2), rgb=C_TOTAL, bold=True, align=A_LEFT_CENTER)
    style(ws.cell(r, 3), bd=B_TB)
    style(ws.cell(r, 4), bd=B_TB)
    style(ws.cell(r, 5), bd=B_RTB)
    ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=5)
    for c in range(first_eff, first_eff + n_eff):
        letter = get_column_letter(c)
        # Summing 3:last would now count every hour at task, group and module level.
        ws.cell(r, c, '=' + '+'.join('%s%d' % (letter, x) for x in l1_rows)
                if l1_rows else '=SUM(%s3:%s%d)' % (letter, letter, last))
        style(ws.cell(r, c), rgb=C_TOTAL, bold=True)
    ws.cell(r, total_col, '=SUM(%s%d:%s%d)' % (eff_a, r, eff_z, r))
    style(ws.cell(r, total_col), rgb=C_TOTAL, bold=True)

    autofit_row_heights(ws, WIDTHS, skip_rows={2})
    ws.row_dimensions[2].height = 36
    return {'name': ws.title, 'spans': spans, 'leaves': leaves, 'total_row': r,
            'total_col': get_column_letter(total_col), 'first_eff': first_eff}


def apply_rollup(ws, first_eff, n_eff, total_col, last):
    """Give every section row a total of its direct children; return the level-1 rows.

    Derived from the sheet that was just written rather than threaded through the writing
    loop, so it works the same for a module that has group rows and for one that carries its
    tasks a level higher.

    Naming each child cell is the whole point. A range would silently include the
    grandchildren as well, which is how the same hours come to be counted at three levels
    once section rows start carrying totals.
    """
    ids = []
    for r in range(3, last + 1):
        t = str(ws.cell(r, 1).value or '').strip()
        if t and t[0].isdigit():
            ids.append((r, t))

    def children_of(idx):
        row, ident = ids[idx]
        depth = ident.count('.')
        kids = []
        for nxt_row, nxt_id in ids[idx + 1:]:
            d = nxt_id.count('.')
            if d <= depth:
                break                      # a sibling or an uncle closes this section
            if d == depth + 1 and nxt_id.startswith(ident + '.'):
                kids.append(nxt_row)
        return kids

    l1_rows = []
    for i, (row, ident) in enumerate(ids):
        if '.' not in ident:
            l1_rows.append(row)
        kids = children_of(i)
        if not kids:
            continue                       # a leaf keeps the value it was given
        for c in range(first_eff, first_eff + n_eff):
            letter = get_column_letter(c)
            ws.cell(row, c, '=' + '+'.join('%s%d' % (letter, k) for k in kids))
            style(ws.cell(row, c), rgb=C_L1 if '.' not in ident else C_L2, bold=True)
        ws.cell(row, total_col,
                '=SUM(%s%d:%s%d)' % (get_column_letter(first_eff), row,
                                     get_column_letter(first_eff + n_eff - 1), row))
        style(ws.cell(row, total_col), rgb=C_L1 if '.' not in ident else C_L2, bold=True)
    return l1_rows


def column_e(item):
    """Assumptions, then the technical risk block when the task carries one.

    The risk lives beside the assumption it qualifies rather than on a separate sheet,
    because that is where the reader is when they ask "why is this number so big".
    """
    text = (item.get('assum') or '').rstrip()
    risk = item.get('risk')
    if risk:
        text += ('\nTechnical risk (%s):\n- %s\n- Mitigation: %s'
                 % (risk['level'], risk['risk'], risk['mitigation']))
    return text


# ------------------------------------------------------------------ Cover
def build_cover(ws, spec, sheets):
    span = 7
    ws.sheet_view.showGridLines = False
    for letter, width in COVER_WIDTHS.items():
        ws.column_dimensions[letter].width = width

    banner(ws, 1, '%s  -  WORK BREAKDOWN STRUCTURE AND EFFORT ESTIMATE'
           % spec['project'].upper(), C_HEADER, span, 14)
    ws.row_dimensions[1].height = 26
    banner(ws, 2, 'Development effort only: development, developer unit testing and '
                  'code-review fixes. Whole hours.', C_EFFORT, span)

    r = 4
    info = [('Project', spec['project']),
            ('Unit', spec.get('currency_note', 'man-hours')),
            ('Mode', 'Structure supplied by the client; effort columns filled'
                     if spec.get('mode') == 'fill'
                     else 'Breakdown authored from the source documents')]
    if spec.get('zero_note'):
        info.append(('Deliberate zeros', spec['zero_note']))
    if spec.get('competitive', 1.0) == 1.0:
        info.append(('Competitive factor', 'Not applied. The hours are the engineering '
                                           'position; any discount is a commercial decision.'))
    for label, value in info:
        ws.cell(r, 1, label)
        style(ws.cell(r, 1), rgb=C_L2, bold=True, align=A_LEFT_CENTER)
        ws.cell(r, 2, value)
        for c in range(2, span + 1):
            style(ws.cell(r, c), align=A_LEFT_CENTER)
        ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=span)
        r += 1

    r += 1
    labels = spec.get('column_labels') or [DEFAULT_LABELS.get(c, c.upper())
                                           for c in spec['columns']]
    for c, h in enumerate(['Sheet / Module'] + labels + ['Total Effort'], start=1):
        ws.cell(r, c, h)
        style(ws.cell(r, c), rgb=C_HEADER if c == 1 else (
            C_HEADER_TOT if c == len(labels) + 2 else C_EFFORT), bold=True, white=True)
    r += 1

    tot_rows = []
    for sh in sheets:
        q = "'%s'!" % sh['name']
        ws.cell(r, 1, sh['name'])
        style(ws.cell(r, 1), rgb=C_L1, bold=True, align=A_LEFT_CENTER)
        for i in range(len(labels) + 1):
            src = get_column_letter(sh['first_eff'] + i) if i < len(labels) else sh['total_col']
            ws.cell(r, 2 + i, '=%s%s%d' % (q, src, sh['total_row']))
            style(ws.cell(r, 2 + i), rgb=C_L1, bold=True)
        tot_rows.append(r)
        r += 1
        for sid, title, a, b in sh['spans']:
            ws.cell(r, 1, '      %s. %s' % (sid, title))
            style(ws.cell(r, 1), align=A_LEFT_CENTER)
            for i in range(len(labels) + 1):
                src = get_column_letter(sh['first_eff'] + i) if i < len(labels) else sh['total_col']
                # `a` is the module's level-1 row, which now carries its own roll-up.
                # Summing a:b would count the module's hours three times over.
                ws.cell(r, 2 + i, '=%s%s%d' % (q, src, a))
                style(ws.cell(r, 2 + i))
            r += 1

    ws.cell(r, 1, 'TOTAL')
    style(ws.cell(r, 1), rgb=C_TOTAL, bold=True, align=A_LEFT_CENTER)
    for i in range(len(labels) + 1):
        letter = get_column_letter(2 + i)
        ws.cell(r, 2 + i, '=' + '+'.join('%s%d' % (letter, x) for x in tot_rows))
        style(ws.cell(r, 2 + i), rgb=C_TOTAL, bold=True)
    r += 2

    lines = [
        'Every estimate covers development, developer unit testing and code-review fixes. '
        'A dedicated QA function, project management and business analysis are not in these '
        'numbers; allow roughly 20-25% and 15-20% respectively on top.',
        'Assumptions that move the hours are recorded per task in column E, including the '
        'assumed COUNT of any external integration. Anything beyond that count is a change '
        'request at the same unit effort.',
        'Tasks carrying real technical risk have a "Technical risk" block in column E with '
        'the risk, its level and the mitigation already built into the approach.',
        'The Out of Scope sheet records what is deliberately excluded and why, so the '
        'boundary of this estimate is explicit rather than inferred.',
    ]
    if spec.get('zero_note'):
        lines.insert(1, spec['zero_note'])
    notes(ws, r, span, 'NOTES', lines)
    autofit_row_heights(ws, COVER_WIDTHS, skip_rows={1})


def build_oos(ws, spec):
    for letter, width in OOS_WIDTHS.items():
        ws.column_dimensions[letter].width = width
    for c, h in enumerate(['No', 'Item', 'Reason for Deferral'], start=1):
        ws.cell(1, c, h)
        style(ws.cell(1, c), rgb=C_OOS_HDR, bold=True)
    for i, entry in enumerate(spec.get('out_of_scope', []), start=1):
        ws.cell(i + 1, 1, i)
        style(ws.cell(i + 1, 1))
        ws.cell(i + 1, 2, entry['item'])
        style(ws.cell(i + 1, 2), align=A_LEFT_CENTER)
        ws.cell(i + 1, 3, entry['reason'])
        style(ws.cell(i + 1, 3), align=A_LEFT_CENTER)
    autofit_row_heights(ws, OOS_WIDTHS)


# ------------------------------------------------------------------ main
def main():
    ap = argparse.ArgumentParser()
    ap.add_argument('--spec', required=True)
    ap.add_argument('--out', required=True)
    a = ap.parse_args()

    spec = wbs_schema.validate(json.load(open(a.spec, encoding='utf-8')))
    log = apply_factors(spec)
    print_factor_log(log, spec.get('competitive', 1.0))
    grand = sum(r.get(c) or 0 for r in wbs_schema.leaf_rows(spec)
                for c in spec['columns'])
    print_competitive_position(spec, grand)

    wb = Workbook()
    cover = wb.active
    cover.title = 'Cover'
    sheets = [build_sheet(wb.create_sheet(s['name'][:31]), spec, set(s['modules']))
              for s in spec['sheets']]
    build_cover(cover, spec, sheets)
    build_oos(wb.create_sheet('Out of Scope'), spec)
    wb.save(a.out)

    cols = spec['columns']
    leaves = wbs_schema.leaf_rows(spec)
    grand = sum(r.get(c) or 0 for r in leaves for c in cols)
    cells = sum(1 for r in leaves for c in cols if r.get(c))
    print('Written: %s' % a.out)
    print('Sheets : %s' % ', '.join(wb.sheetnames))
    print()
    print('%-34s %7s %9s' % ('Sheet', 'tasks', 'hours'))
    for sh in sheets:
        h = sum(r.get(c) or 0 for r in sh['leaves'] for c in cols)
        print('%-34s %7d %9d' % (sh['name'][:34], len(sh['leaves']), h))
    print('%-34s %7d %9d' % ('TOTAL', len(leaves), grand))
    print()
    if cells:
        print('Average per leaf task       : %.1f h' % (grand / len(leaves)))
        print('Average per populated cell  : %.1f h  (%d cells)' % (grand / cells, cells))
    for c in cols:
        s = sum(r.get(c) or 0 for r in leaves)
        print('  %-5s %7d  %5.1f%%' % (c, s, 100.0 * s / grand if grand else 0))


if __name__ == '__main__':
    main()
