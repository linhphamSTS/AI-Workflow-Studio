# -*- coding: utf-8 -*-
"""Shared Excel styling for the WBS and cost workbooks.

The colour scheme matches the estimate workbooks this skill was derived from, so the
two deliverables read as one family.

The row-height code is the important part. openpyxl writes no row height at all, and
Excel Online and SharePoint do NOT recalculate one: they render each row at its stored
height, which falls back to a single line, so every wrapped cell is clipped and the
reader has to drag each row open by hand. Excel Desktop hides the problem by
recalculating on open, which is why a reference workbook can look fine locally and
still be broken for the client.

The per-line figure is not a guess. Heights that Excel itself stored in a reference
workbook solve exactly to `n * 15.0 + 0.75` for Calibri 11, so a line is a flat 15.0 pt
and Excel allows only 0.75 pt of slack. That is why even a correctly auto-fitted cell
reads tight against its border, and why this module adds real bottom padding instead.
"""
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

# --------------------------------------------------------------------- palette
C_HEADER = 'FF1C4587'      # ID / Module / Feature / Description / Assumptions
C_HEADER_TOT = 'FF1C4586'  # the total column, one shade off in the reference file
C_EFFORT = 'FF38761D'      # the effort band
C_L1 = 'FFFCE4D6'          # level-1 section row
C_L2 = 'FFFFF2CC'          # level-2 group row
C_LEAF = 'FFFFFFFF'        # leaf task row
C_TOTAL = 'FFFFFF00'       # TOTAL row
C_OOS_HDR = 'FFD9EAD3'     # Out of Scope header

FONT = 'Calibri'
SZ = 11.0
MONEY = '#,##0.00'
PCT = '0.0%'

_THIN = Side(style='thin')


def border(spec):
    """`border('LRTB')` -> a Border with those edges."""
    return Border(
        left=_THIN if 'L' in spec else None,
        right=_THIN if 'R' in spec else None,
        top=_THIN if 'T' in spec else None,
        bottom=_THIN if 'B' in spec else None,
    )


B_ALL = border('LRTB')
B_TB = border('TB')
B_RTB = border('RTB')
B_LTB = border('LTB')
B_NONE = Border()

A_CENTER = Alignment(horizontal='center', vertical='center', wrap_text=True)
A_LEFT_TOP = Alignment(horizontal='left', vertical='top', wrap_text=True)
A_LEFT_CENTER = Alignment(horizontal='left', vertical='center', wrap_text=True)


def fill(rgb):
    return PatternFill(fill_type='solid', fgColor=rgb)


def style(cell, *, rgb=None, bold=False, white=False, align=A_CENTER, bd=B_ALL, size=SZ):
    cell.font = Font(name=FONT, size=size, bold=bold,
                     color='FFFFFFFF' if white else 'FF000000')
    if rgb:
        cell.fill = fill(rgb)
    cell.alignment = align
    cell.border = bd
    return cell


def banner(ws, row, text, rgb, span, size=11):
    """A full-width coloured band, used for section titles."""
    ws.cell(row, 1, text)
    style(ws.cell(row, 1), rgb=rgb, bold=True, white=True, align=A_LEFT_CENTER, size=size)
    for col in range(2, span + 1):
        style(ws.cell(row, col), rgb=rgb)
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=span)


def notes(ws, row, span, title, lines):
    """A titled block of borderless bullet lines. Returns the next free row."""
    banner(ws, row, title, C_HEADER, span)
    row += 1
    for text in lines:
        ws.cell(row, 1, '-  ' + text)
        style(ws.cell(row, 1), align=A_LEFT_TOP, bd=B_NONE)
        for col in range(2, span + 1):
            style(ws.cell(row, col), bd=B_NONE)
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=span)
        row += 1
    return row


# ----------------------------------------------------------------- row heights
PT_PER_LINE = 15.0        # Calibri 11, every line. 14.4 under-sizes every row.
BOTTOM_PAD_PT = 6.0       # about a third of a line, more generous than Excel's 0.75
MAX_ROW_PT = 409.0        # Excel's ceiling
CELL_PAD_PX = 10          # left+right padding, generous so wrapping is never under-counted


def _measure_font():
    """Calibri 11 at 96 DPI is about 15 px. Falls back to a metric guess."""
    try:
        from PIL import ImageFont
        for name in ('calibri.ttf', 'Calibri.ttf', 'arial.ttf', 'DejaVuSans.ttf'):
            try:
                return ImageFont.truetype(name, 15)
            except OSError:
                continue
    except ImportError:
        pass
    return None


_FONT = _measure_font()


def _text_px(s):
    if _FONT is not None:
        return _FONT.getlength(s)
    return len(s) * 7.0


def _col_px(width):
    """Excel column width (in '0' characters) -> usable pixels."""
    return width * 7.0 + 5.0 - CELL_PAD_PX


def wrapped_lines(text, avail_px):
    """How many lines the text occupies in a cell that wide, honouring line breaks."""
    if not text:
        return 1
    total = 0
    for para in str(text).split('\n'):
        if not para.strip():
            total += 1
            continue
        if _text_px(para) <= avail_px:
            total += 1
            continue
        line, count = '', 1
        for word in para.split(' '):
            probe = word if not line else line + ' ' + word
            if _text_px(probe) <= avail_px:
                line = probe
                continue
            if line:
                count += 1
                line = word
            else:                          # a single word wider than the cell
                count += max(1, int(_text_px(word) / avail_px))
                line = ''
        total += count
    return max(1, total)


def autofit_row_heights(ws, widths, skip_rows=()):
    """Stamp an explicit height on every row so wrapped text is fully visible.

    Rows listed in `skip_rows` are left alone; give them a height yourself, or they
    render as one line on SharePoint.
    """
    spans, vmerge = {}, {}
    for rng in ws.merged_cells.ranges:
        if rng.max_col > rng.min_col:
            spans[(rng.min_row, rng.min_col)] = sum(
                widths.get(get_column_letter(c), 8.43)
                for c in range(rng.min_col, rng.max_col + 1))
        if rng.max_row > rng.min_row:
            vmerge[(rng.min_row, rng.min_col)] = rng.max_row - rng.min_row + 1

    for row in range(1, ws.max_row + 1):
        if row in skip_rows:
            continue
        lines = 1
        for col in range(1, ws.max_column + 1):
            cell = ws.cell(row, col)
            v = cell.value
            if not isinstance(v, str) or v.startswith('='):
                continue
            if not (cell.alignment and cell.alignment.wrap_text):
                continue
            width = spans.get((row, col)) or widths.get(get_column_letter(col), 8.43)
            need = wrapped_lines(v, _col_px(width))
            down = vmerge.get((row, col), 1)
            if down > 1:
                need = -(-need // down)          # a merged cell shares its text down the rows
            lines = max(lines, need)
        ws.row_dimensions[row].height = min(
            MAX_ROW_PT, round(lines * PT_PER_LINE + BOTTOM_PAD_PT, 2))


def header_fits(ws, widths, header_rows=(1, 2)):
    """Return the header cells whose text needs more height than their row provides.

    A column left at Excel's default width once cut the tail off its own header, and the
    defect was inherited from the reference workbook where nobody had noticed it.
    """
    bad = []
    avail_total = sum((ws.row_dimensions[r].height or 15.0) for r in header_rows)
    hspan = {}
    for rng in ws.merged_cells.ranges:
        if rng.max_col > rng.min_col and rng.min_row <= max(header_rows):
            hspan[(rng.min_row, rng.min_col)] = sum(
                widths.get(get_column_letter(c), 8.43)
                for c in range(rng.min_col, rng.max_col + 1))
    for row in header_rows:
        avail = avail_total if row == min(header_rows) else (
            ws.row_dimensions[row].height or 15.0)
        for col in range(1, ws.max_column + 1):
            v = ws.cell(row, col).value
            if not isinstance(v, str) or not v.strip():
                continue
            width = hspan.get((row, col)) or widths.get(get_column_letter(col), 8.43)
            need = wrapped_lines(v, _col_px(width)) * PT_PER_LINE
            if need > avail + 0.5:
                bad.append('%s%d %r needs %.0fpt, has %.0fpt'
                           % (get_column_letter(col), row, v[:24], need, avail))
    return bad
