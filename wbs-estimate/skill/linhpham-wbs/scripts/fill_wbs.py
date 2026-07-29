#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""Fill hours into a WBS the client supplied, and change nothing else.

    python fill_wbs.py --spec wbs.json --client "WBS from client.xlsx" \\
                       --out "WBS_Acme.xlsx"

This is the common case: the client has already written the breakdown and wants numbers in
it. Their structure is the deliverable, so it is not rebuilt, reordered, restyled or
re-worded. The effort columns are located by their own headers, rows are matched by task
identifier, and only those cells are written.

Two refusals, both deliberate:

* if a task in the estimate has no row in their workbook the fill STOPS. Writing the rest
  would deliver a workbook that is short by however many tasks failed to match, and the
  total would look plausible.
* if a row in their workbook is a task and the estimate has nothing for it, the fill stops
  as well. An unpriced task in a priced workbook reads as free.

Row heights are stamped on the way out unless told not to. Excel Online and SharePoint never
auto-fit, so a wrapped cell in their file clips for their reviewer even though it looked
correct on the author's desktop.
"""
import argparse
import json
import os
import re
import sys

import openpyxl

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
import wbs_schema  # noqa: E402

# Header text that identifies an effort column, per canonical column key. Matched loosely
# because a client writes "BE (h)", "Back-end", "Backend Dev" and all three mean the same.
HEADER_HINTS = {
    'ui':  ('ui/ux', 'ui / ux', 'uiux', 'design'),
    'be':  ('be', 'back-end', 'back end', 'backend', 'server'),
    'fe':  ('fe', 'front-end', 'front end', 'frontend', 'web'),
    'mob': ('mobile', 'ios', 'android', 'app'),
    'ai':  ('ai', 'artificial'),
    'devops': ('devops', 'dev ops', 'infra', 'infrastructure'),
    'qa':  ('qa', 'test', 'quality'),
}
ID_HINTS = ('id', 'wbs', 'no', 'item', 'task id', 'ref')


def norm(v):
    return re.sub(r'\s+', ' ', str(v or '')).strip().lower()


def find_layout(ws, cols, max_scan=25):
    """Locate the header row, the id column and one column per effort key."""
    for r in range(1, min(ws.max_row, max_scan) + 1):
        texts = {c: norm(ws.cell(r, c).value) for c in range(1, ws.max_column + 1)}
        if not any(texts.values()):
            continue
        found = {}
        for key in cols:
            hints = HEADER_HINTS.get(key, (key,))
            for c, t in texts.items():
                if not t:
                    continue
                if t in hints or any(t.startswith(h) or h == t for h in hints) \
                        or any(re.fullmatch(r'%s\s*\(?h(ours?)?\)?' % re.escape(h), t)
                               for h in hints):
                    found.setdefault(key, c)
                    break
        if len(found) < max(1, len(cols) // 2):
            continue
        id_col = None
        for c, t in texts.items():
            if t in ID_HINTS or any(t.startswith(h) for h in ID_HINTS):
                id_col = c
                break
        if id_col is None:
            id_col = 1
        return r, id_col, found
    return None, None, {}


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument('--spec', required=True)
    ap.add_argument('--client', required=True, help="the client's workbook")
    ap.add_argument('--out', required=True)
    ap.add_argument('--sheet', help='limit to one sheet; default is every sheet that has a '
                                    'recognisable header row')
    ap.add_argument('--no-row-heights', action='store_true')
    a = ap.parse_args()

    spec = wbs_schema.validate(json.load(open(a.spec, encoding='utf-8')))
    cols = spec['columns']
    want = {}
    for leaf in wbs_schema.leaf_rows(spec):
        want[str(leaf['id']).strip()] = leaf

    # Load the client file and save to the output path only once everything checks out.
    # Copying first meant a refused fill still left a file behind, so the message said
    # nothing was saved while a workbook sat there looking like a deliverable.
    wb = openpyxl.load_workbook(a.client)

    sheets = [a.sheet] if a.sheet else list(wb.sheetnames)
    written, unmatched_rows, touched_sheets = set(), [], []

    for name in sheets:
        ws = wb[name]
        hdr, id_col, found = find_layout(ws, cols)
        if not hdr or not found:
            continue
        touched_sheets.append((name, hdr, id_col, found))
        for r in range(hdr + 1, ws.max_row + 1):
            rid = str(ws.cell(r, id_col).value or '').strip()
            if not rid or not rid[0].isdigit():
                continue
            if '.' not in rid:
                continue                     # a section row: their headers stay untouched
            leaf = want.get(rid)
            if leaf is None:
                # A task row in their file that the estimate has nothing for.
                if any(ws.cell(r, c).value not in (None, '') for c in found.values()):
                    continue                 # already carries numbers of their own
                unmatched_rows.append('%s!%s' % (name, rid))
                continue
            for key, c in found.items():
                v = leaf.get(key)
                ws.cell(r, c, v if v else None)
            written.add(rid)

    if not touched_sheets:
        raise SystemExit(
            'No sheet in %s had a header row naming the effort columns %s.\n'
            'Name the columns in the spec the way the client names them, or pass --sheet.'
            % (os.path.basename(a.client), ', '.join(cols)))

    missing = sorted(set(want) - written, key=lambda x: [int(p) for p in x.split('.')
                                                         if p.isdigit()])
    problems = []
    if missing:
        problems.append('%d estimated task(s) had no row in the client workbook, so their '
                        'hours were not delivered: %s'
                        % (len(missing), ', '.join(missing[:12])))
    if unmatched_rows:
        problems.append('%d task row(s) in the client workbook have no estimate and no hours '
                        'of their own, which reads as free: %s'
                        % (len(unmatched_rows), ', '.join(unmatched_rows[:12])))

    print('=' * 78)
    print('FILL  -  %s' % os.path.basename(a.client))
    print('=' * 78)
    for name, hdr, id_col, found in touched_sheets:
        print('  %-28s header row %-3d id col %-3d  %s'
              % (name[:28], hdr, id_col,
                 ', '.join('%s=%s' % (k, openpyxl.utils.get_column_letter(c))
                           for k, c in sorted(found.items()))))
    print('  filled %d of %d estimated task(s)' % (len(written), len(want)))

    if problems:
        print()
        for p in problems:
            print('  ! ' + p)
        print()
        print('  Nothing was saved. A partial fill produces a total that looks plausible and '
              'is short, which is the one outcome worse than an error.')
        return 1

    if not a.no_row_heights:
        from xlsx_style import autofit_row_heights
        for name, _hdr, _idc, _f in touched_sheets:
            ws = wb[name]
            widths = {openpyxl.utils.get_column_letter(c):
                      (ws.column_dimensions[openpyxl.utils.get_column_letter(c)].width or 8.43)
                      for c in range(1, ws.max_column + 1)}
            try:
                autofit_row_heights(ws, widths)
            except Exception as exc:
                print('  row heights on %s skipped: %s' % (name, exc))

    wb.save(a.out)
    total = sum(leaf.get(c) or 0 for leaf in want.values() for c in cols)
    print('  total %d h' % total)
    print('  wrote %s' % a.out)
    return 0


if __name__ == '__main__':
    sys.exit(main())
