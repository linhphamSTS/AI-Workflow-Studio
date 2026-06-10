#!/usr/bin/env python3
"""Read an .xlsx workbook and emit one Markdown-table-per-sheet so the
Phase-1 ingest agent can analyse RFP requirement tables, scoring matrices,
WBS spreadsheets, etc."""
from __future__ import annotations

import argparse
import sys
from pathlib import Path


def _ensure_openpyxl():
    try:
        import openpyxl  # noqa: F401
    except ImportError:
        import subprocess
        subprocess.run(
            [sys.executable, "-m", "pip", "install", "--quiet", "--user", "openpyxl"],
            check=True,
        )


def read(path: Path) -> str:
    _ensure_openpyxl()
    from openpyxl import load_workbook
    wb = load_workbook(str(path), data_only=True, read_only=True)
    out = [f"# Workbook: {path.name}", ""]
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        out.append(f"## Sheet: {sheet_name}")
        rows = list(ws.iter_rows(values_only=True))
        # Trim trailing empty rows/cols
        while rows and all(c is None or str(c).strip() == "" for c in rows[-1]):
            rows.pop()
        if not rows:
            out.append("(empty sheet)\n")
            continue
        # Render as a markdown table; treat first non-empty row as header.
        max_cols = max(len(r) for r in rows)
        norm = [list(r) + [None] * (max_cols - len(r)) for r in rows]
        header = ["" if c is None else str(c).replace("|", "/").strip() for c in norm[0]]
        out.append("| " + " | ".join(header) + " |")
        out.append("|" + "|".join("---" for _ in header) + "|")
        for r in norm[1:]:
            cells = ["" if c is None else str(c).replace("|", "/").replace("\n", " ").strip() for c in r]
            out.append("| " + " | ".join(cells) + " |")
        out.append("")
    return "\n".join(out)


def main() -> int:
    ap = argparse.ArgumentParser()
    ap.add_argument("path", type=Path)
    args = ap.parse_args()
    if not args.path.exists():
        print(f"! file not found: {args.path}", file=sys.stderr)
        return 1
    sys.stdout.write(read(args.path))
    return 0


if __name__ == "__main__":
    sys.exit(main())
